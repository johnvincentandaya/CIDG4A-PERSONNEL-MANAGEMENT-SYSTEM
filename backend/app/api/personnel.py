from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from typing import List, Optional
from sqlalchemy import func
from sqlalchemy.orm import Session
from .. import models, schemas
from ..database import SessionLocal, init_db
from ..utils import ensure_upload_folders, personnel_folder_name, uploads_abs, uploads_rel
import os
from datetime import datetime
from io import BytesIO
import hashlib
import json
import openpyxl.styles
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from fastapi.responses import StreamingResponse
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch

router = APIRouter()

def _upper_str(val):
    if val is None:
        return None
    if isinstance(val, str):
        return val.upper()
    return val

ALLOWED_FILE_EXTENSIONS = {".pdf", ".doc", ".docx"}
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024


def abs_from_stored_path(stored_path: Optional[str]):
    if not stored_path:
        return None
    norm = str(stored_path).replace("\\", "/")
    idx = norm.find("uploads/")
    rel = norm[idx + len("uploads/"):] if idx >= 0 else norm.lstrip("/")
    return str(uploads_abs(*[p for p in rel.split("/") if p]))


async def read_and_validate_upload(file_obj: UploadFile):
    ext = (os.path.splitext(file_obj.filename or "")[1] or "").lower()
    if ext not in ALLOWED_FILE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type for '{file_obj.filename}'. Allowed types: PDF, DOC, DOCX.",
        )
    content = await file_obj.read()
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise HTTPException(
            status_code=400,
            detail=f"File '{file_obj.filename}' exceeds 10MB size limit.",
        )
    return ext, content


def file_sha256(path: Optional[str]):
    if not path or not os.path.exists(path):
        return None
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(8192)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def validate_personnel_inputs(
    rank: Optional[str],
    status: Optional[str],
    status_custom: Optional[str],
    nup_rank: Optional[str],
    nup_entry_number: Optional[int],
):
    rank_upper = _upper_str(rank) or ''
    status_upper = _upper_str(status) or ''

    if status_upper == 'OTHERS' and not str(status_custom or '').strip():
        raise HTTPException(status_code=400, detail='Status details are required when status is Others.')

    if rank_upper == 'NUP':
        if not str(nup_rank or '').strip():
            raise HTTPException(status_code=400, detail='NUP rank is required when rank is NUP.')
        if nup_entry_number is None:
            raise HTTPException(status_code=400, detail='NUP entry number is required when rank is NUP.')


def validate_required_profile_fields(parsed_birthdate, religion: Optional[str]):
    if parsed_birthdate is None:
        raise HTTPException(status_code=400, detail='Birthdate is required and must be a valid date (YYYY-MM-DD).')
    if not str(religion or '').strip():
        raise HTTPException(status_code=400, detail='Religion is required.')


@router.on_event("startup")
def startup():
    ensure_upload_folders()
    init_db()


@router.post("/basic", response_model=schemas.PersonnelSchema)
def create_personnel_basic(
    rank: str = Form(...),
    badge_number: Optional[str] = Form(None),
    last_name: str = Form(...),
    first_name: str = Form(...),
    mi: Optional[str] = Form(None),
    suffix: Optional[str] = Form(None),
    unit: str = Form(...),
    status: str = Form(...),
    status_custom: Optional[str] = Form(None),
    qlf: Optional[str] = Form(None),
    nup_rank: Optional[str] = Form(None),
    nup_entry_number: Optional[int] = Form(None),
    date_of_reassignment: Optional[str] = Form(None),
    designation: Optional[str] = Form(None),
    date_of_designation: Optional[str] = Form(None),
    highest_eligibility: Optional[str] = Form(None),
    contact_number: Optional[str] = Form(None),
    birthdate: Optional[str] = Form(None),
    religion: Optional[str] = Form(None),
    section: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """
    Create a personnel record without requiring any documents.
    Documents and trainings can be attached later via the update endpoint.
    """
    def _parse_date(s: Optional[str]):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            try:
                return datetime.fromisoformat(s)
            except Exception:
                return None

    parsed_reassignment = _parse_date(date_of_reassignment)
    parsed_designation_date = _parse_date(date_of_designation)
    parsed_birthdate = _parse_date(birthdate)
    validate_required_profile_fields(parsed_birthdate, religion)

    validate_personnel_inputs(rank, status, status_custom, nup_rank, nup_entry_number)

    p = models.Personnel(
        rank=_upper_str(rank),
        badge_number=_upper_str(badge_number),
        last_name=_upper_str(last_name),
        first_name=_upper_str(first_name),
        mi=_upper_str(mi),
        suffix=_upper_str(suffix),
        unit=_upper_str(unit),
        status=_upper_str(status),
        status_custom=_upper_str(status_custom),
        qlf=_upper_str(qlf),
        nup_rank=_upper_str(nup_rank),
        nup_entry_number=nup_entry_number,
        date_of_reassignment=parsed_reassignment,
        designation=_upper_str(designation),
        date_of_designation=parsed_designation_date,
        highest_eligibility=_upper_str(highest_eligibility),
        contact_number=_upper_str(contact_number),
        birthdate=parsed_birthdate,
        religion=_upper_str(religion),
        section=_upper_str(section),
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return p


@router.post("/", response_model=schemas.PersonnelSchema)
async def create_personnel(
    rank: str = Form(...),
    badge_number: Optional[str] = Form(None),
    last_name: str = Form(...),
    first_name: str = Form(...),
    mi: Optional[str] = Form(None),
    suffix: Optional[str] = Form(None),
    unit: str = Form(...),
    status: str = Form(...),
    status_custom: Optional[str] = Form(None),
    qlf: Optional[str] = Form(None),
    nup_rank: Optional[str] = Form(None),
    nup_entry_number: Optional[int] = Form(None),
    date_of_reassignment: Optional[str] = Form(None),
    designation: Optional[str] = Form(None),
    date_of_designation: Optional[str] = Form(None),
    highest_eligibility: Optional[str] = Form(None),
    contact_number: Optional[str] = Form(None),
    birthdate: Optional[str] = Form(None),
    religion: Optional[str] = Form(None),
    section: Optional[str] = Form(None),
    # single-file documents (now optional so record can be saved even if incomplete)
    pds: Optional[UploadFile] = File(None),
    appointment: Optional[UploadFile] = File(None),
    promotion: Optional[UploadFile] = File(None),
    designation_file: Optional[UploadFile] = File(None),
    reassignment: Optional[UploadFile] = File(None),
    diploma: Optional[UploadFile] = File(None),
    eligibility: Optional[UploadFile] = File(None),
    iper: Optional[UploadFile] = File(None),
    saln: Optional[UploadFile] = File(None),
    pft: Optional[UploadFile] = File(None),
    rca: Optional[UploadFile] = File(None),
    # trainings: lists (optional)
    mandatory_files: Optional[List[UploadFile]] = File(None),
    mandatory_titles: Optional[List[str]] = Form(None),
    specialized_files: Optional[List[UploadFile]] = File(None),
    specialized_titles: Optional[List[str]] = Form(None),
    db: Session = Depends(get_db),
):

    validate_personnel_inputs(rank, status, status_custom, nup_rank, nup_entry_number)

    # parse date strings into datetimes where provided
    def _parse_date(s: Optional[str]):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            try:
                return datetime.fromisoformat(s)
            except Exception:
                return None

    parsed_reassignment = _parse_date(date_of_reassignment)
    parsed_designation_date = _parse_date(date_of_designation)
    parsed_birthdate = _parse_date(birthdate)
    validate_required_profile_fields(parsed_birthdate, religion)

    # Normalize textual inputs to uppercase before saving and before using for filenames
    rank = _upper_str(rank)
    badge_number = _upper_str(badge_number)
    last_name = _upper_str(last_name)
    first_name = _upper_str(first_name)
    mi = _upper_str(mi)
    suffix = _upper_str(suffix)
    unit = _upper_str(unit)
    status = _upper_str(status)
    status_custom = _upper_str(status_custom)
    qlf = _upper_str(qlf)
    nup_rank = _upper_str(nup_rank)
    designation = _upper_str(designation)
    highest_eligibility = _upper_str(highest_eligibility)
    contact_number = _upper_str(contact_number)
    religion = _upper_str(religion)
    section = _upper_str(section)

    # create personnel entry (badge_number is optional)
    p = models.Personnel(
        rank=rank,
        badge_number=badge_number,
        last_name=last_name,
        first_name=first_name,
        mi=mi,
        suffix=suffix,
        unit=unit,
        status=status,
        status_custom=status_custom,
        qlf=qlf,
        nup_rank=nup_rank,
        nup_entry_number=nup_entry_number,
        designation=designation,
        highest_eligibility=highest_eligibility,
        contact_number=contact_number,
        birthdate=parsed_birthdate,
        religion=religion,
        section=section,
        date_of_reassignment=parsed_reassignment,
        date_of_designation=parsed_designation_date,
    )
    db.add(p)
    db.commit()
    db.refresh(p)

    # save uploaded files into structured folder (by unit) and record them
    unit_folder = unit.upper()
    person_folder_abs = uploads_abs('form_201', unit_folder, personnel_folder_name(first_name, last_name))
    person_folder_rel = uploads_rel('form_201', unit_folder, personnel_folder_name(first_name, last_name))
    os.makedirs(person_folder_abs, exist_ok=True)

    async def save_single(file_obj: Optional[UploadFile], shortname: str):
        if not file_obj:
            return
        ext, content = await read_and_validate_upload(file_obj)
        # Use required file naming: FORM201_<First>_<Last>_<document>.pdf
        dest_name = f"FORM201_{first_name}_{last_name}_{shortname}{ext}"
        dest_path_abs = os.path.join(str(person_folder_abs), dest_name)
        with open(dest_path_abs, 'wb') as out:
            out.write(content)
        rel_path = f"{person_folder_rel}/{dest_name}".replace("\\", "/")
        doc = models.Document(personnel_id=p.id, doc_type=shortname, file_path=rel_path)
        db.add(doc)

    # map and save
    await save_single(pds, "pds")
    await save_single(appointment, "appointment")
    await save_single(promotion, "promotion")
    await save_single(designation_file, "designation")
    await save_single(reassignment, "reassignment")
    await save_single(diploma, "diploma")
    await save_single(eligibility, "eligibility")
    await save_single(iper, "iper")
    await save_single(saln, "saln")
    await save_single(pft, "pft")
    await save_single(rca, "rca")

    # trainings
    # save mandatory
    if mandatory_files:
        for idx, f in enumerate(mandatory_files):
            title = ""
            if mandatory_titles and idx < len(mandatory_titles):
                title = mandatory_titles[idx]
            safe_title = title.replace(" ", "_") if title else f"mandatory_{idx+1}"
            ext, content = await read_and_validate_upload(f)
            dest_name = f"FORM201_{first_name}_{last_name}_mandatory_{safe_title}{ext}"
            dest_path_abs = os.path.join(str(person_folder_abs), dest_name)
            with open(dest_path_abs, "wb") as out:
                out.write(content)
            norm_path = f"{person_folder_rel}/{dest_name}".replace("\\", "/")
            tc = models.TrainingCertificate(
                personnel_id=p.id,
                category="mandatory",
                title=(title.upper() if title else safe_title.upper()),
                file_path=norm_path,
            )
            db.add(tc)

    # save specialized
    if specialized_files:
        for idx, f in enumerate(specialized_files):
            title = ""
            if specialized_titles and idx < len(specialized_titles):
                title = specialized_titles[idx]
            safe_title = title.replace(" ", "_") if title else f"specialized_{idx+1}"
            ext, content = await read_and_validate_upload(f)
            dest_name = f"FORM201_{first_name}_{last_name}_specialized_{safe_title}{ext}"
            dest_path_abs = os.path.join(str(person_folder_abs), dest_name)
            with open(dest_path_abs, "wb") as out:
                out.write(content)
            norm_path = f"{person_folder_rel}/{dest_name}".replace("\\", "/")
            tc = models.TrainingCertificate(
                personnel_id=p.id,
                category="specialized",
                title=(title.upper() if title else safe_title.upper()),
                file_path=norm_path,
            )
            db.add(tc)

    db.commit()
    db.refresh(p)
    return p


@router.get("/", response_model=List[schemas.PersonnelSchema])
def list_personnel(db: Session = Depends(get_db)):
    return db.query(models.Personnel).order_by(models.Personnel.last_name).all()


@router.get("/counts")
def personnel_counts(db: Session = Depends(get_db)):
    from sqlalchemy import func

    unit_map = {
        'RHQ': 'RHQ',
        'Cavite': 'CAVITE',
        'Laguna': 'LAGUNA',
        'Batangas': 'BATANGAS',
        'Rizal': 'RIZAL',
        'Quezon': 'QUEZON',
    }
    counts = {
        display_name: db.query(models.Personnel).filter(func.upper(models.Personnel.unit) == db_unit).count()
        for display_name, db_unit in unit_map.items()
    }
    counts['total'] = sum(counts.values())
    return counts


@router.get("/{person_id}/documents")
def get_documents(person_id: int, db: Session = Depends(get_db)):
    person = db.query(models.Personnel).filter(models.Personnel.id==person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail='Personnel not found')
    return person.documents


@router.get("/{person_id}", response_model=schemas.PersonnelSchema)
def get_person(person_id: int, db: Session = Depends(get_db)):
    person = db.query(models.Personnel).filter(models.Personnel.id==person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail='Personnel not found')
    return person


@router.put("/{person_id}", response_model=schemas.PersonnelSchema)
async def update_person(
    person_id: int,
    rank: str = Form(...),
    badge_number: Optional[str] = Form(None),
    last_name: str = Form(...),
    first_name: str = Form(...),
    mi: Optional[str] = Form(None),
    suffix: Optional[str] = Form(None),
    unit: str = Form(...),
    status: str = Form(...),
    status_custom: Optional[str] = Form(None),
    qlf: Optional[str] = Form(None),
    nup_rank: Optional[str] = Form(None),
    nup_entry_number: Optional[int] = Form(None),
    date_of_reassignment: Optional[str] = Form(None),
    designation: Optional[str] = Form(None),
    date_of_designation: Optional[str] = Form(None),
    highest_eligibility: Optional[str] = Form(None),
    contact_number: Optional[str] = Form(None),
    birthdate: Optional[str] = Form(None),
    religion: Optional[str] = Form(None),
    section: Optional[str] = Form(None),
    # optional replacement single files
    pds: Optional[UploadFile] = File(None),
    appointment: Optional[UploadFile] = File(None),
    promotion: Optional[UploadFile] = File(None),
    designation_file: Optional[UploadFile] = File(None),
    reassignment: Optional[UploadFile] = File(None),
    diploma: Optional[UploadFile] = File(None),
    eligibility: Optional[UploadFile] = File(None),
    iper: Optional[UploadFile] = File(None),
    saln: Optional[UploadFile] = File(None),
    pft: Optional[UploadFile] = File(None),
    rca: Optional[UploadFile] = File(None),
    # trainings/documents: deletions and additions
    delete_doc_types: Optional[str] = Form(None),
    delete_mandatory_ids: Optional[str] = Form(None),
    delete_specialized_ids: Optional[str] = Form(None),
    mandatory_files: Optional[List[UploadFile]] = File(None),
    mandatory_titles: Optional[List[str]] = Form(None),
    specialized_files: Optional[List[UploadFile]] = File(None),
    specialized_titles: Optional[List[str]] = Form(None),
    db: Session = Depends(get_db),
):
    validate_personnel_inputs(rank, status, status_custom, nup_rank, nup_entry_number)

    person = db.query(models.Personnel).filter(models.Personnel.id==person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail='Personnel not found')

    # update basic fields
    # parse incoming date strings
    def _parse_date(s: Optional[str]):
        if not s:
            return None
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            try:
                return datetime.fromisoformat(s)
            except Exception:
                return None

    parsed_birthdate = _parse_date(birthdate)
    validate_required_profile_fields(parsed_birthdate, religion)

    # Normalize incoming textual values to uppercase before saving
    person.rank = _upper_str(rank)
    person.badge_number = _upper_str(badge_number)
    person.last_name = _upper_str(last_name)
    person.first_name = _upper_str(first_name)
    person.mi = _upper_str(mi)
    person.suffix = _upper_str(suffix)
    person.unit = _upper_str(unit)
    person.status = _upper_str(status)
    person.status_custom = _upper_str(status_custom)
    person.qlf = _upper_str(qlf)
    person.nup_rank = _upper_str(nup_rank)
    person.nup_entry_number = nup_entry_number
    person.designation = _upper_str(designation)
    person.date_of_reassignment = _parse_date(date_of_reassignment)
    person.date_of_designation = _parse_date(date_of_designation)
    person.highest_eligibility = _upper_str(highest_eligibility)
    person.contact_number = _upper_str(contact_number)
    person.birthdate = parsed_birthdate
    person.religion = _upper_str(religion)
    person.section = _upper_str(section)
    db.add(person)
    db.commit()

    # person folder (by unit)
    unit_folder = unit.upper()
    person_folder_abs = uploads_abs('form_201', unit_folder, personnel_folder_name(first_name, last_name))
    person_folder_rel = uploads_rel('form_201', unit_folder, personnel_folder_name(first_name, last_name))
    os.makedirs(person_folder_abs, exist_ok=True)

    # helper to replace or create document
    async def replace_single(file_obj: UploadFile, shortname: str):
        if not file_obj:
            return
        ext, content = await read_and_validate_upload(file_obj)
        dest_name = f"FORM201_{first_name}_{last_name}_{shortname}{ext}"
        dest_path_abs = os.path.join(str(person_folder_abs), dest_name)
        norm_path = f"{person_folder_rel}/{dest_name}".replace("\\", "/")
        # check existing doc
        doc = db.query(models.Document).filter(models.Document.personnel_id==person.id, models.Document.doc_type==shortname).first()
        existing_abs = abs_from_stored_path(doc.file_path) if doc else None
        if existing_abs and os.path.exists(existing_abs):
            old_hash = file_sha256(existing_abs)
            new_hash = hashlib.sha256(content).hexdigest()
            if old_hash == new_hash:
                return
        with open(dest_path_abs, 'wb') as out:
            out.write(content)
        if doc:
            if existing_abs and os.path.exists(existing_abs) and os.path.abspath(existing_abs) != os.path.abspath(dest_path_abs):
                try:
                    os.remove(existing_abs)
                except OSError:
                    pass
            doc.file_path = norm_path
            db.add(doc)
        else:
            doc = models.Document(personnel_id=person.id, doc_type=shortname, file_path=norm_path)
            db.add(doc)

    await replace_single(pds, 'pds')
    await replace_single(appointment, 'appointment')
    await replace_single(promotion, 'promotion')
    await replace_single(designation_file, 'designation')
    await replace_single(reassignment, 'reassignment')
    await replace_single(diploma, 'diploma')
    await replace_single(eligibility, 'eligibility')
    await replace_single(iper, 'iper')
    await replace_single(saln, 'saln')
    await replace_single(pft, 'pft')
    await replace_single(rca, 'rca')

    # handle deletion of existing single documents
    if delete_doc_types:
        try:
            doc_types = json.loads(delete_doc_types)
            for d_type in doc_types:
                doc = db.query(models.Document).filter(models.Document.personnel_id==person.id, models.Document.doc_type==d_type).first()
                if doc:
                    abs_path = abs_from_stored_path(doc.file_path)
                    if abs_path and os.path.exists(abs_path):
                        try:
                            os.remove(abs_path)
                        except OSError:
                            pass
                    db.delete(doc)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail='Invalid delete_doc_types payload.')

    # handle deletions of trainings
    if delete_mandatory_ids:
        try:
            ids = json.loads(delete_mandatory_ids)
            for iid in ids:
                t = db.query(models.TrainingCertificate).filter(models.TrainingCertificate.id==iid, models.TrainingCertificate.personnel_id==person.id).first()
                if t:
                    # remove file
                    try:
                        abs_path = abs_from_stored_path(t.file_path)
                        if abs_path and os.path.exists(abs_path):
                            os.remove(abs_path)
                    except Exception:
                        pass
                    db.delete(t)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail='Invalid delete_mandatory_ids payload.')
    if delete_specialized_ids:
        try:
            ids = json.loads(delete_specialized_ids)
            for iid in ids:
                t = db.query(models.TrainingCertificate).filter(models.TrainingCertificate.id==iid, models.TrainingCertificate.personnel_id==person.id).first()
                if t:
                    try:
                        abs_path = abs_from_stored_path(t.file_path)
                        if abs_path and os.path.exists(abs_path):
                            os.remove(abs_path)
                    except Exception:
                        pass
                    db.delete(t)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail='Invalid delete_specialized_ids payload.')

    # add new trainings if any
    if mandatory_files:
        for idx, f in enumerate(mandatory_files):
            title = ''
            if mandatory_titles and idx < len(mandatory_titles):
                title = mandatory_titles[idx]
            safe_title = title.replace(' ', '_') if title else f'mandatory_new_{idx+1}'
            ext, content = await read_and_validate_upload(f)
            dest_name = f"FORM201_{first_name}_{last_name}_mandatory_{safe_title}{ext}"
            dest_path_abs = os.path.join(str(person_folder_abs), dest_name)
            with open(dest_path_abs, 'wb') as out:
                out.write(content)
            norm_path = f"{person_folder_rel}/{dest_name}".replace("\\", "/")
            tc = models.TrainingCertificate(personnel_id=person.id, category='mandatory', title=(title.upper() if title else safe_title.upper()), file_path=norm_path)
            db.add(tc)

    if specialized_files:
        for idx, f in enumerate(specialized_files):
            title = ''
            if specialized_titles and idx < len(specialized_titles):
                title = specialized_titles[idx]
            safe_title = title.replace(' ', '_') if title else f'specialized_new_{idx+1}'
            ext, content = await read_and_validate_upload(f)
            dest_name = f"FORM201_{first_name}_{last_name}_specialized_{safe_title}{ext}"
            dest_path_abs = os.path.join(str(person_folder_abs), dest_name)
            with open(dest_path_abs, 'wb') as out:
                out.write(content)
            norm_path = f"{person_folder_rel}/{dest_name}".replace("\\", "/")
            tc = models.TrainingCertificate(personnel_id=person.id, category='specialized', title=(title.upper() if title else safe_title.upper()), file_path=norm_path)
            db.add(tc)

    db.commit()
    db.refresh(person)
    return person


@router.delete("/{person_id}")
def delete_person(person_id: int, db: Session = Depends(get_db)):
    person = db.query(models.Personnel).filter(models.Personnel.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail='Personnel not found')

    file_paths = []
    for doc in (person.documents or []):
        abs_path = abs_from_stored_path(doc.file_path)
        if abs_path:
            file_paths.append(abs_path)

    for tr in (person.trainings or []):
        abs_path = abs_from_stored_path(tr.file_path)
        if abs_path:
            file_paths.append(abs_path)

    for bmi in (person.bmi_records or []):
        for photo_path in [bmi.photo_front, bmi.photo_left, bmi.photo_right]:
            abs_path = abs_from_stored_path(photo_path)
            if abs_path:
                file_paths.append(abs_path)

    db.delete(person)
    db.commit()

    for fp in file_paths:
        try:
            if fp and os.path.exists(fp):
                os.remove(fp)
        except OSError:
            pass

    return {"message": "Personnel record deleted successfully"}


@router.post('/form201-report')
def generate_form201_report(
    unit: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    prepared_by: str = Form(''),
    noted_by: str = Form(''),
    file_name: str = Form('form201_report'),
    db: Session = Depends(get_db),
):
    """
    Generate a Form 201 Personnel Report grouped by unit.
    
    Document Completion Logic:
    - 11 single document types + 1 mandatory training + 1 specialized training = 13 total
    - Completion: count_uploaded / 13
    - Status: "COMPLETE" if all 13, else "WITH DEFICIENCY"
    """
    from openpyxl.styles import Font, Border, Side, Alignment
    
    # Sanitize filename
    safe_name = (file_name or "form201_report").strip()
    safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in safe_name)
    safe_name = safe_name.strip("_") or "form201_report"
    
    # Query personnel with filters
    query = db.query(models.Personnel)
    if unit and unit != 'All Units':
        query = query.filter(models.Personnel.unit == unit)
    if status and status != 'All Status':
        query = query.filter(models.Personnel.status == status)
    
    all_personnel = query.all()
    
    # Group by unit and sort
    units = ['RHQ', 'CAVITE', 'LAGUNA', 'BATANGAS', 'RIZAL', 'QUEZON']
    grouped = {u: [] for u in units}
    for person in all_personnel:
        unit_key = person.unit.upper() if person.unit else 'RHQ'
        if unit_key in grouped:
            grouped[unit_key].append(person)
    
    # Sort each unit: by rank (if available), then by surname
    def rank_order(rank_str: str) -> int:
        rank_hierarchy = {
            'PGen': 1, 'MGen': 2, 'BGen': 3, 'Col': 4, 'LtCol': 5, 'Maj': 6,
            'Cpt': 7, '1st Lt': 8, '2nd Lt': 9, 'Spy': 10, 'PSP': 11, 'PFC': 12,
        }
        return rank_hierarchy.get(rank_str or '', 999)
    
    for unit_key in grouped:
        grouped[unit_key].sort(key=lambda p: (
            rank_order(p.rank),
            (p.last_name or '').lower(),
        ))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Form 201 Report'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    
    current_row = 1
    
    # Header
    ws.merge_cells(f'A{current_row}:I{current_row}')
    header = ws[f'A{current_row}']
    header.value = 'CRIMINAL INVESTIGATION AND DETECTION GROUP REGION 4A'
    header.font = Font(bold=True, size=12)
    header.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:I{current_row}')
    subheader = ws[f'A{current_row}']
    subheader.value = 'FORM 201 PERSONNEL REPORT'
    subheader.font = Font(bold=True, size=11)
    subheader.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:I{current_row}')
    date_cell = ws[f'A{current_row}']
    date_cell.value = f"(as of {datetime.utcnow().strftime('%B %d, %Y')})"
    date_cell.font = Font(italic=True, size=10)
    date_cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2
    
    # Process each unit
    for unit_key in units:
        personnel_list = grouped[unit_key]
        if not personnel_list:
            continue
        
        # Unit header
        ws.merge_cells(f'A{current_row}:I{current_row}')
        unit_cell = ws[f'A{current_row}']
        unit_cell.value = f'UNIT: {unit_key}'
        unit_cell.font = Font(bold=True, size=11)
        unit_cell.fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        unit_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        # Table header
        headers = ['No.', 'Rank', 'Surname', 'First Name', 'Middle Name', 'Suffix', 'Status', 'Document Completion', 'Remarks']
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Personnel rows
        for idx, person in enumerate(personnel_list, start=1):
            single_doc_count = len(person.documents or [])
            has_mandatory = any(t.category == 'mandatory' for t in (person.trainings or []))
            has_specialized = any(t.category == 'specialized' for t in (person.trainings or []))
            
            total_count = single_doc_count
            if has_mandatory:
                total_count += 1
            if has_specialized:
                total_count += 1
            
            completion_str = f"{total_count}/13"
            is_complete = total_count == 13
            remarks = "COMPLETE" if is_complete else "WITH DEFICIENCY"
            
            # Determine missing documents for remarks (optional)
            missing = []
            single_doc_types = ['pds', 'appointment', 'promotion', 'designation', 'reassignment', 'diploma', 'eligibility', 'iper', 'saln', 'pft', 'rca']
            existing_types = {d.doc_type for d in (person.documents or [])}
            for doc_type in single_doc_types:
                if doc_type not in existing_types:
                    missing.append(doc_type.upper())
            if not has_mandatory:
                missing.append('MANDATORY TRAINING')
            if not has_specialized:
                missing.append('SPECIALIZED TRAINING')
            
            row_data = [
                idx,
                person.rank or '',
                person.last_name or '',
                person.first_name or '',
                person.mi or '',
                person.suffix or '',
                person.status or '',
                completion_str,
                remarks,
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        current_row += 1
    
    # Footer section
    current_row += 1
    ws[f'A{current_row}'] = 'Prepared by: ________________________'
    ws[f'A{current_row}'].font = Font(name='Courier New', size=10)
    current_row += 1
    ws[f'D{current_row}'] = 'Noted by: ________________________'
    ws[f'D{current_row}'].font = Font(name='Courier New', size=10)
    
    # If prepared_by or noted_by provided, add them
    if prepared_by or noted_by:
        current_row += 3
        if prepared_by:
            ws[f'A{current_row}'] = prepared_by
        if noted_by:
            ws[f'D{current_row}'] = noted_by
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 20
    
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment;filename={safe_name}.xlsx"},
    )


@router.post('/report')
def personnel_report(
    unit: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    report_type: str = Form('excel'),
    file_name: str = Form('form201_report'),
    # New parameters from Generate Modal
    as_of_date: Optional[str] = Form(None),
    scope: Optional[str] = Form(None),
    specific_unit: Optional[str] = Form(None),
    # Signatories
    prepared_by_name: Optional[str] = Form(None),
    prepared_by_title: Optional[str] = Form(None),
    prepared_by_signature: Optional[UploadFile] = File(None),
    verified_by_name: Optional[str] = Form(None),
    verified_by_title: Optional[str] = Form(None),
    verified_by_signature: Optional[UploadFile] = File(None),
    noted_by_name: Optional[str] = Form(None),
    noted_by_title: Optional[str] = Form(None),
    noted_by_signature: Optional[UploadFile] = File(None),
    # Authorized values from modal
    authorized_values: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    def to_int(v, default=0):
        try:
            if v is None or v == '':
                return default
            return int(v)
        except Exception:
            return default

    def safe(v):
        if v is None:
            return ''
        return str(v).strip()

    def up(v):
        return safe(v).upper()

    def fmt_date(v):
        if not v:
            return ''
        try:
            return v.strftime('%Y-%m-%d')
        except Exception:
            return str(v)

    aut_values = {'PCO': {}, 'PNCO': {}, 'NUP': {'NUP': 0}, 'Totals': {}}
    if authorized_values:
        try:
            parsed = json.loads(authorized_values)
            if isinstance(parsed, dict):
                aut_values = parsed
        except Exception:
            pass

    pco_aut = aut_values.get('PCO') if isinstance(aut_values.get('PCO'), dict) else {}
    pnco_aut = aut_values.get('PNCO') if isinstance(aut_values.get('PNCO'), dict) else {}
    nup_aut_raw = aut_values.get('NUP', {})
    if isinstance(nup_aut_raw, dict):
        nup_aut = to_int(nup_aut_raw.get('NUP', 0), 0)
    else:
        nup_aut = to_int(nup_aut_raw, 0)
    totals_aut = aut_values.get('Totals') if isinstance(aut_values.get('Totals'), dict) else {}

    report_date = as_of_date or datetime.utcnow().strftime('%Y-%m-%d')
    try:
        display_date = datetime.strptime(report_date, '%Y-%m-%d').strftime('%B %d, %Y')
    except Exception:
        display_date = datetime.utcnow().strftime('%B %d, %Y')

    def unit_key(raw_unit: str):
        u = up(raw_unit)
        if u in ('RHQ', 'RFU4A HQS', 'REGIONAL OFFICE', 'HEADQUARTERS'):
            return 'RHQ'
        if 'CAV' in u:
            return 'CAVITE'
        if 'LAG' in u:
            return 'LAGUNA'
        if 'BAT' in u:
            return 'BATANGAS'
        if 'RIZ' in u:
            return 'RIZAL'
        if 'QZN' in u or 'QUE' in u:
            return 'QUEZON'
        return u

    def unit_variants(raw_unit: Optional[str]):
        key = unit_key(raw_unit)
        variants = {
            'RHQ': ['RHQ', 'RFU4A HQS', 'REGIONAL OFFICE', 'HEADQUARTERS'],
            'CAVITE': ['CAVITE', 'CAVITE PFU', 'CAV PFU'],
            'LAGUNA': ['LAGUNA', 'LAGUNA PFU', 'LAG PFU'],
            'BATANGAS': ['BATANGAS', 'BATANGAS PFU', 'BATS PFU'],
            'RIZAL': ['RIZAL', 'RIZAL PFU'],
            'QUEZON': ['QUEZON', 'QUEZON PFU', 'QZN PFU'],
        }
        return variants.get(key, [key] if key else [])

    q = db.query(models.Personnel)
    if scope == 'RHQ only':
        q = q.filter(func.upper(models.Personnel.unit).in_(unit_variants('RHQ')))
    elif scope == 'Specific Unit' and specific_unit:
        unit_opts = unit_variants(specific_unit)
        if unit_opts:
            q = q.filter(func.upper(models.Personnel.unit).in_(unit_opts))
    elif unit and unit != 'All Units':
        unit_opts = unit_variants(unit)
        if unit_opts:
            q = q.filter(func.upper(models.Personnel.unit).in_(unit_opts))
    if status and status != 'All Status':
        q = q.filter(func.upper(models.Personnel.status) == up(status))
    all_personnel = q.all()

    PCO_RANKS = ['PMGEN', 'PBGEN', 'PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT']
    PNCO_RANKS = ['PEMS', 'PCMS', 'PSMS', 'PMSG', 'PSSG', 'PCPL', 'PAT']
    UNIFORMED_ORDER = ['PGEN', 'PLTGEN', 'PMGEN', 'PBGEN', 'PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT', 'PEMS', 'PCMS', 'PSMS', 'PMSG', 'PSSG', 'PCPL', 'PAT']
    UNIFORMED_ORDER_MAP = {r: i for i, r in enumerate(UNIFORMED_ORDER)}

    def is_nup(p):
        return up(getattr(p, 'rank', '')) == 'NUP' or up(getattr(p, 'status', '')) == 'NUP'

    def person_rank(p):
        if is_nup(p):
            return up(getattr(p, 'nup_rank', '') or getattr(p, 'rank', 'NUP'))
        return up(getattr(p, 'rank', ''))

    def is_pco(p):
        return person_rank(p) in PCO_RANKS

    def is_pnco(p):
        return person_rank(p) in PNCO_RANKS

    def rank_sort_key(p):
        r = person_rank(p)
        return (UNIFORMED_ORDER_MAP.get(r, 999), up(p.last_name), up(p.first_name), up(p.mi), up(p.suffix))

    def alpha_sort_key(p):
        return (up(p.last_name), up(p.first_name), up(p.mi), up(p.suffix), person_rank(p))

    def entry_no_sort_key(p):
        return (to_int(getattr(p, 'nup_entry_number', None), 999999), up(p.last_name), up(p.first_name))

    signature_present = {
        'prepared': prepared_by_signature is not None,
        'verified': verified_by_signature is not None,
        'noted': noted_by_signature is not None,
    }

    wb = Workbook()
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    section_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    def write_title(ws, title, cols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        ws.cell(row=1, column=1).value = 'CRIMINAL INVESTIGATION AND DETECTION GROUP REGION 4A'
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        ws.cell(row=2, column=1).value = title
        ws.cell(row=2, column=1).font = Font(bold=True, size=11)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cols)
        ws.cell(row=3, column=1).value = f'(AS OF {display_date.upper()})'
        ws.cell(row=3, column=1).font = Font(italic=True, size=10)
        ws.cell(row=3, column=1).alignment = Alignment(horizontal='center')

    def write_header(ws, row, headers):
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i)
            c.value = h
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin

    def write_row(ws, row, values, center_from=2):
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i)
            c.value = v
            c.border = thin
            c.alignment = Alignment(horizontal='center' if i >= center_from else 'left', vertical='center')

    def write_signatures(ws, start_row, checked_label='Verified Correct by:'):
        blocks = [
            ('Prepared by:', prepared_by_name, prepared_by_title, signature_present['prepared']),
            (checked_label, verified_by_name, verified_by_title, signature_present['verified']),
            ('Noted by:', noted_by_name, noted_by_title, signature_present['noted']),
        ]
        max_col = max(ws.max_column, 9)
        base = max_col // 3
        rem = max_col % 3
        spans = []
        col_start = 1
        for idx in range(3):
            width = base + (1 if idx < rem else 0)
            col_end = col_start + max(width, 1) - 1
            spans.append((col_start, col_end))
            col_start = col_end + 1

        for idx, (label, name, title, esig) in enumerate(blocks):
            c1, c2 = spans[idx]
            ws.merge_cells(start_row=start_row, start_column=c1, end_row=start_row, end_column=c2)
            ws.merge_cells(start_row=start_row + 1, start_column=c1, end_row=start_row + 1, end_column=c2)
            ws.merge_cells(start_row=start_row + 2, start_column=c1, end_row=start_row + 2, end_column=c2)
            ws.merge_cells(start_row=start_row + 3, start_column=c1, end_row=start_row + 3, end_column=c2)
            ws.merge_cells(start_row=start_row + 4, start_column=c1, end_row=start_row + 4, end_column=c2)

            ws.cell(row=start_row, column=c1).value = label
            ws.cell(row=start_row, column=c1).font = Font(bold=True)
            ws.cell(row=start_row, column=c1).alignment = Alignment(horizontal='center')

            ws.cell(row=start_row + 1, column=c1).value = 'E-SIGN: ON FILE' if esig else ''
            ws.cell(row=start_row + 1, column=c1).alignment = Alignment(horizontal='center')

            ws.cell(row=start_row + 2, column=c1).value = up(name)
            ws.cell(row=start_row + 2, column=c1).font = Font(bold=True)
            ws.cell(row=start_row + 2, column=c1).alignment = Alignment(horizontal='center')

            ws.cell(row=start_row + 3, column=c1).value = up(title)
            ws.cell(row=start_row + 3, column=c1).alignment = Alignment(horizontal='center')

            ws.cell(row=start_row + 4, column=c1).value = ''
            ws.cell(row=start_row + 4, column=c1).alignment = Alignment(horizontal='center')

        ws.row_dimensions[start_row + 1].height = 24

    # TAB 1: LIST OF NUP
    ws = wb.active
    ws.title = 'LIST OF NUP'
    write_title(ws, 'LIST OF NON-UNIFORMED PERSONNEL', 7)
    headers = ['Entry No.', 'Rank', 'Last', 'First', 'Middle', 'Q', 'Office']
    row = 5
    write_header(ws, row, headers)
    row += 1
    nup_personnel = sorted([p for p in all_personnel if is_nup(p)], key=entry_no_sort_key)
    for idx, p in enumerate(nup_personnel, start=1):
        write_row(ws, row, [
            to_int(p.nup_entry_number, idx),
            person_rank(p),
            up(p.last_name),
            up(p.first_name),
            up(p.mi),
            up(p.suffix),
            up(p.unit),
        ])
        row += 1
    write_signatures(ws, row + 2)
    for i, w in enumerate([10, 12, 20, 20, 14, 8, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 2: LIST OF U.P
    ws = wb.create_sheet('LIST OF U.P')
    write_title(ws, 'LIST OF UNIFORMED PERSONNEL', 8)
    headers = ['Entry No.', 'Rank', 'Last', 'First', 'Middle', 'Q', 'Badge No.', 'Office']
    row = 5
    write_header(ws, row, headers)
    row += 1
    up_personnel = sorted([p for p in all_personnel if not is_nup(p)], key=rank_sort_key)
    for idx, p in enumerate(up_personnel, start=1):
        write_row(ws, row, [idx, person_rank(p), up(p.last_name), up(p.first_name), up(p.mi), up(p.suffix), up(p.badge_number), up(p.unit)])
        row += 1
    write_signatures(ws, row + 2)
    for i, w in enumerate([10, 12, 20, 20, 14, 8, 14, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 3: TERRITORIAL STRENGTH
    ws = wb.create_sheet('TERRITORIAL STRENGTH')
    write_title(ws, 'TERRITORIAL STRENGTH', 5)
    row = 5
    write_header(ws, row, ['Office/Unit', 'PCOs', 'PNCOs', 'NUPs', 'Total'])
    row += 1
    terr_rows = [
        ('RHQ', 'RFU4A HQS'),
        ('CAVITE', 'CAVITE PFU'),
        ('LAGUNA', 'LAGUNA PFU'),
        ('BATANGAS', 'BATANGAS PFU'),
        ('RIZAL', 'RIZAL PFU'),
        ('QUEZON', 'QZN PFU'),
    ]
    t_pco = t_pnco = t_nup = 0
    for key, label in terr_rows:
        members = [p for p in all_personnel if unit_key(p.unit) == key]
        pco_count = sum(1 for p in members if is_pco(p))
        pnco_count = sum(1 for p in members if is_pnco(p))
        nup_count = sum(1 for p in members if is_nup(p))
        total_count = pco_count + pnco_count + nup_count
        write_row(ws, row, [label, pco_count, pnco_count, nup_count, total_count])
        row += 1
        t_pco += pco_count
        t_pnco += pnco_count
        t_nup += nup_count
    write_row(ws, row, ['GRAND TOTAL', t_pco, t_pnco, t_nup, t_pco + t_pnco + t_nup])
    for c in range(1, 6):
        ws.cell(row=row, column=c).font = Font(bold=True)
    write_signatures(ws, row + 2)
    for i, w in enumerate([22, 10, 10, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 4: RANK STRUCTURE (no signatures)
    ws = wb.create_sheet('RANK STRUCTURE')
    write_title(ws, 'RANK STRUCTURE', 6)
    row = 5
    write_header(ws, row, ['Rank', 'AUT', 'ACT', 'VAR'])
    row += 1

    pco_actual = {r: sum(1 for p in all_personnel if person_rank(p) == r) for r in PCO_RANKS}
    pnco_actual = {r: sum(1 for p in all_personnel if person_rank(p) == r) for r in PNCO_RANKS}
    for r in PCO_RANKS:
        aut = to_int(pco_aut.get(r), 0)
        act = pco_actual.get(r, 0)
        write_row(ws, row, [r, aut, act, act - aut])
        row += 1
    pco_aut_sub = to_int(pco_aut.get('Subtotal'), sum(to_int(pco_aut.get(r), 0) for r in PCO_RANKS))
    pco_act_sub = sum(pco_actual.values())
    write_row(ws, row, ['Sub-Total (PCOs)', pco_aut_sub, pco_act_sub, pco_act_sub - pco_aut_sub], center_from=1)
    for c in range(1, 5):
        ws.cell(row=row, column=c).font = Font(bold=True)
    row += 1
    for r in PNCO_RANKS:
        aut = to_int(pnco_aut.get(r), 0)
        act = pnco_actual.get(r, 0)
        write_row(ws, row, [r, aut, act, act - aut])
        row += 1
    pnco_aut_sub = to_int(pnco_aut.get('Subtotal'), sum(to_int(pnco_aut.get(r), 0) for r in PNCO_RANKS))
    pnco_act_sub = sum(pnco_actual.values())
    write_row(ws, row, ['Sub-Total (PNCOs)', pnco_aut_sub, pnco_act_sub, pnco_act_sub - pnco_aut_sub], center_from=1)
    for c in range(1, 5):
        ws.cell(row=row, column=c).font = Font(bold=True)
    row += 1
    nup_act = sum(1 for p in all_personnel if is_nup(p))
    write_row(ws, row, ['NUP', nup_aut, nup_act, nup_act - nup_aut], center_from=1)
    row += 1
    total_uniformed_aut = to_int(totals_aut.get('TotalUniformed', pco_aut_sub + pnco_aut_sub), pco_aut_sub + pnco_aut_sub)
    total_uniformed_act = pco_act_sub + pnco_act_sub
    write_row(ws, row, ['Total Uniformed Personnel', total_uniformed_aut, total_uniformed_act, total_uniformed_act - total_uniformed_aut], center_from=1)
    for c in range(1, 5):
        ws.cell(row=row, column=c).font = Font(bold=True)
    row += 2
    total_personnel_aut = to_int(totals_aut.get('TotalPersonnel', total_uniformed_aut + nup_aut), total_uniformed_aut + nup_aut)
    total_personnel_act = total_uniformed_act + nup_act
    ws.cell(row=row, column=1).value = 'ACTUAL STRENGTH'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = total_personnel_act
    ws.cell(row=row + 1, column=1).value = 'AUTHORIZED STRENGTH'
    ws.cell(row=row + 1, column=1).font = Font(bold=True)
    ws.cell(row=row + 1, column=2).value = total_personnel_aut
    for i, w in enumerate([24, 12, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 5: RANK PROFILE
    ws = wb.create_sheet('RANK PROFILE')
    write_title(ws, 'RANK PROFILE', 18)
    row = 5
    headers = ['Unit', 'PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT', 'Actual PCO', 'PEMS', 'PCMS', 'PSMS', 'PMSG', 'PSSG', 'PCPL', 'PAT', 'Actual PNCO', 'Sub-Total Uniformed Personnel', 'Actual NUP', 'Total']
    write_header(ws, row, headers)
    row += 1
    profile_units = [('RHQ', 'RFU4A HQS'), ('CAVITE', 'CAV PFU'), ('LAGUNA', 'LAG PFU'), ('BATANGAS', 'BATS PFU'), ('RIZAL', 'RIZAL PFU'), ('QUEZON', 'QZN PFU')]
    totals = [0] * (len(headers) - 1)
    for key, label in profile_units:
        members = [p for p in all_personnel if unit_key(p.unit) == key]
        pco_counts = [sum(1 for p in members if person_rank(p) == r) for r in ['PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT']]
        pnco_counts = [sum(1 for p in members if person_rank(p) == r) for r in PNCO_RANKS]
        actual_pco = sum(pco_counts)
        actual_pnco = sum(pnco_counts)
        uniformed = actual_pco + actual_pnco
        actual_nup = sum(1 for p in members if is_nup(p))
        total = uniformed + actual_nup
        values = pco_counts + [actual_pco] + pnco_counts + [actual_pnco, uniformed, actual_nup, total]
        write_row(ws, row, [label] + values)
        totals = [a + b for a, b in zip(totals, values)]
        row += 1
    write_row(ws, row, ['TOTAL'] + totals)
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = section_fill
    write_signatures(ws, row + 2)
    for i, w in enumerate([16, 7, 8, 7, 7, 7, 10, 7, 7, 7, 7, 7, 7, 7, 10, 14, 10, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 6: ALPHA LIST
    ws = wb.create_sheet('ALPHA LIST')
    write_title(ws, 'ALPHA LIST', 12)
    row = 5
    headers = ['Entry Number', 'Account Number', 'Rank', 'Last', 'First', 'Middle', 'Q', 'Badge No.', 'Contact Number', 'Office', 'Date of Birth', 'Religion']
    write_header(ws, row, headers)
    row += 1
    alpha_rows = sorted(all_personnel, key=alpha_sort_key)
    for idx, p in enumerate(alpha_rows, start=1):
        entry_number = to_int(p.nup_entry_number, idx) if is_nup(p) else idx
        write_row(ws, row, [entry_number, up(p.badge_number), person_rank(p), up(p.last_name), up(p.first_name), up(p.mi), up(p.suffix), up(p.badge_number), safe(p.contact_number), up(p.unit), fmt_date(p.birthdate), up(p.religion)])
        row += 1
    write_signatures(ws, row + 2)
    for i, w in enumerate([12, 14, 10, 18, 18, 12, 8, 12, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 7: DISPOSITION OF TROOPS
    ws = wb.create_sheet('DISPOSITION OF TROOPS')
    write_title(ws, 'DISPOSITION OF TROOPS', 16)
    row = 5
    headers = ['No.', 'Badge No.', 'Rank', 'Last', 'First', 'Middle', 'Qualifier', 'Date of Reassignment', 'Designation', 'Date of Designation', 'Mandatory Training', 'Specialized Training', 'Highest Eligibility', 'Contact Number', 'Birthdate', 'Religion']
    group_order = [
        ('Office of the Regional Chief', 'RHQ_SECTION'),
        ('Admin and HRDD Section', 'RHQ_SECTION'),
        ('Intelligence Section', 'RHQ_SECTION'),
        ('Investigation Section', 'RHQ_SECTION'),
        ('Operation & PCR Section', 'RHQ_SECTION'),
        ('Cavite', 'CAVITE'),
        ('Laguna', 'LAGUNA'),
        ('Batangas', 'BATANGAS'),
        ('Rizal', 'RIZAL'),
        ('Quezon', 'QUEZON'),
    ]

    def in_group(p, display_name, group_type):
        if group_type == 'RHQ_SECTION':
            if unit_key(p.unit) != 'RHQ':
                return False
            section_name = safe(p.section)
            if not section_name:
                section_name = 'Office of the Regional Chief'
            return up(section_name) == up(display_name)
        return unit_key(p.unit) == group_type

    dispo_pco = dispo_pnco = dispo_nup = 0
    for display_name, group_type in group_order:
        members = [p for p in all_personnel if in_group(p, display_name, group_type)]
        if not members:
            continue
        members = sorted(members, key=rank_sort_key)
        ws.cell(row=row, column=1).value = up(display_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1
        write_header(ws, row, headers)
        row += 1
        for idx, p in enumerate(members, start=1):
            mandatory = '; '.join([up(t.title) for t in (p.trainings or []) if up(t.category) == 'MANDATORY'])
            specialized = '; '.join([up(t.title) for t in (p.trainings or []) if up(t.category) == 'SPECIALIZED'])
            write_row(ws, row, [
                idx,
                up(p.badge_number),
                person_rank(p),
                up(p.last_name),
                up(p.first_name),
                up(p.mi),
                up(p.suffix),
                fmt_date(p.date_of_reassignment),
                up(p.designation),
                fmt_date(p.date_of_designation),
                mandatory,
                specialized,
                up(p.highest_eligibility),
                safe(p.contact_number),
                fmt_date(p.birthdate),
                up(p.religion),
            ])
            row += 1
        dispo_pco += sum(1 for p in members if is_pco(p))
        dispo_pnco += sum(1 for p in members if is_pnco(p))
        dispo_nup += sum(1 for p in members if is_nup(p))
        row += 1

    write_header(ws, row, ['', 'ACT', 'AUT', 'VAR'])
    row += 1
    summary_rows = [
        ('PCO', dispo_pco, pco_aut_sub),
        ('PNCO', dispo_pnco, pnco_aut_sub),
        ('NUP', dispo_nup, nup_aut),
        ('TOTAL', dispo_pco + dispo_pnco + dispo_nup, pco_aut_sub + pnco_aut_sub + nup_aut),
    ]
    for label, act, aut in summary_rows:
        write_row(ws, row, [label, act, aut, act - aut], center_from=1)
        for c in range(1, 5):
            ws.cell(row=row, column=c).font = Font(bold=True if label == 'TOTAL' else False)
        row += 1
    write_signatures(ws, row + 1, checked_label='Checked by:')
    for i, w in enumerate([8, 12, 10, 16, 16, 12, 10, 14, 20, 14, 24, 24, 18, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 8: PERSONNEL FILL-UP
    ws = wb.create_sheet('PERSONNEL FILL-UP')
    write_title(ws, 'PERSONNEL FILL-UP', 17)
    row = 5
    headers = ['Office', 'PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT', 'TOTAL', 'PEMS', 'PCMS', 'PSMS', 'PMSG', 'PSSG', 'PCPL', 'PAT', 'TOTAL', 'NUP', 'GRAND TOTAL']
    write_header(ws, row, headers)
    row += 1
    fill_units = [('RHQ', 'Regional Office'), ('CAVITE', 'Cavite'), ('LAGUNA', 'Laguna'), ('BATANGAS', 'Batangas'), ('RIZAL', 'Rizal'), ('QUEZON', 'Quezon')]
    running_totals = [0] * 16
    for key, label in fill_units:
        members = [p for p in all_personnel if unit_key(p.unit) == key]
        pco_counts = [sum(1 for p in members if person_rank(p) == r) for r in ['PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT']]
        pnco_counts = [sum(1 for p in members if person_rank(p) == r) for r in PNCO_RANKS]
        pco_total = sum(pco_counts)
        pnco_total = sum(pnco_counts)
        nup_count = sum(1 for p in members if is_nup(p))
        grand = pco_total + pnco_total + nup_count
        values = pco_counts + [pco_total] + pnco_counts + [pnco_total, nup_count, grand]
        write_row(ws, row, [label] + values)
        running_totals = [a + b for a, b in zip(running_totals, values)]
        row += 1
    write_row(ws, row, ['TOTAL'] + running_totals)
    for c in range(1, 18):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = section_fill
    write_signatures(ws, row + 2)
    for i, w in enumerate([16, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 9: KEY OFFICERS
    ws = wb.create_sheet('KEY OFFICERS')
    write_title(ws, 'KEY OFFICERS', 6)
    ws.cell(row=4, column=1).value = 'CIDG REGION 4A'
    ws.cell(row=4, column=1).font = Font(bold=True)
    row = 5
    write_header(ws, row, ['Designation', 'Authorized Rank', 'Rank', 'Name', 'Date of Designation', 'Contact Number'])
    row += 1
    key_officers = sorted([p for p in all_personnel if safe(p.designation)], key=lambda p: (up(p.designation), rank_sort_key(p)))
    for p in key_officers:
        officer_name = f"{up(p.last_name)}, {up(p.first_name)} {up(p.mi)}".strip()
        write_row(ws, row, [up(p.designation), person_rank(p), person_rank(p), officer_name, fmt_date(p.date_of_designation), safe(p.contact_number)], center_from=2)
        row += 1
    write_signatures(ws, row + 2)
    for i, w in enumerate([24, 16, 12, 30, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 10: STATION LIST
    ws = wb.create_sheet('STATION LIST')
    write_title(ws, 'STATION LIST', 9)
    row = 5
    write_header(ws, row, ['Badge Number', 'SCOMM', 'Rank', 'Last', 'First', 'Middle', 'Qualifier', 'Date of Designation', 'Designation'])
    row += 1
    station_rows = sorted([p for p in all_personnel if safe(p.designation)], key=rank_sort_key)
    for p in station_rows:
        write_row(ws, row, [up(p.badge_number), up(p.badge_number), person_rank(p), up(p.last_name), up(p.first_name), up(p.mi), up(p.suffix), fmt_date(p.date_of_designation), up(p.designation)])
        row += 1
    row += 1
    ws.cell(row=row, column=1).value = 'RECAPITULATION'
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    recap_counts = [
        ('PCOL', sum(1 for p in station_rows if person_rank(p) == 'PCOL')),
        ('PLTCOL', sum(1 for p in station_rows if person_rank(p) == 'PLTCOL')),
        ('PMAJ', sum(1 for p in station_rows if person_rank(p) == 'PMAJ')),
        ('PCPT', sum(1 for p in station_rows if person_rank(p) == 'PCPT')),
        ('PLT', sum(1 for p in station_rows if person_rank(p) == 'PLT')),
    ]
    for label, count in recap_counts:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = count
        row += 1
    ws.cell(row=row, column=1).value = 'TOTAL'
    ws.cell(row=row, column=2).value = sum(count for _, count in recap_counts)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    write_signatures(ws, row + 2)
    for i, w in enumerate([14, 14, 10, 18, 18, 12, 10, 16, 25], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # TAB 11: RANK INVENTORY
    ws = wb.create_sheet('RANK INVENTORY')
    write_title(ws, 'RANK INVENTORY', 2)
    row = 5
    ws.cell(row=row, column=1).value = 'RFU'
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = 'CIDG REGION 4A'
    row += 1
    ws.cell(row=row, column=1).value = 'Date'
    ws.cell(row=row, column=2).value = display_date.upper()
    row += 2
    ws.cell(row=row, column=1).value = 'PCO'
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    pco_total = 0
    for rank in ['PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT']:
        count = sum(1 for p in all_personnel if person_rank(p) == rank)
        pco_total += count
        ws.cell(row=row, column=1).value = rank
        ws.cell(row=row, column=2).value = count
        row += 1
    ws.cell(row=row, column=1).value = 'Total'
    ws.cell(row=row, column=2).value = pco_total
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 2
    ws.cell(row=row, column=1).value = 'PNCO'
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    pnco_total = 0
    for rank in PNCO_RANKS:
        count = sum(1 for p in all_personnel if person_rank(p) == rank)
        pnco_total += count
        ws.cell(row=row, column=1).value = rank
        ws.cell(row=row, column=2).value = count
        row += 1
    ws.cell(row=row, column=1).value = 'Total'
    ws.cell(row=row, column=2).value = pnco_total
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 2
    ws.cell(row=row, column=1).value = 'NUP'
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    nup_total = sum(1 for p in all_personnel if is_nup(p))
    ws.cell(row=row, column=1).value = 'NUP'
    ws.cell(row=row, column=2).value = nup_total
    row += 1
    ws.cell(row=row, column=1).value = 'Total'
    ws.cell(row=row, column=2).value = nup_total
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 2
    ws.cell(row=row, column=1).value = 'Total Strength'
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1).value = 'PCO total'
    ws.cell(row=row, column=2).value = pco_total
    row += 1
    ws.cell(row=row, column=1).value = 'PNCO total'
    ws.cell(row=row, column=2).value = pnco_total
    row += 1
    ws.cell(row=row, column=1).value = 'NUP total'
    ws.cell(row=row, column=2).value = nup_total
    row += 1
    ws.cell(row=row, column=1).value = 'Grand Total'
    ws.cell(row=row, column=2).value = pco_total + pnco_total + nup_total
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14

    # TAB 12: CIDG RFU4A LIST OF PCOs AND CONTACT NUMBERS
    ws = wb.create_sheet('CIDG RFU4A PCO CONTACT LIST')
    write_title(ws, 'CIDG RFU4A LIST OF PCOs AND CONTACT NUMBERS', 9)
    row = 5
    write_header(ws, row, ['Badge No.', 'SCOMM', 'Rank', 'Last', 'First', 'MI', 'Qualifier', 'Designation', 'Contact Number'])
    row += 1
    office_order = [('RHQ', 'Regional Office'), ('CAVITE', 'Cavite PFU'), ('LAGUNA', 'Laguna PFU'), ('BATANGAS', 'Batangas PFU'), ('RIZAL', 'Rizal PFU'), ('QUEZON', 'Quezon PFU')]
    for office_key, office_label in office_order:
        members = [p for p in all_personnel if unit_key(p.unit) == office_key and is_pco(p)]
        if not members:
            continue
        members = sorted(members, key=rank_sort_key)
        ws.cell(row=row, column=1).value = up(office_label)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        for p in members:
            write_row(ws, row, [up(p.badge_number), up(p.badge_number), person_rank(p), up(p.last_name), up(p.first_name), up(p.mi), up(p.suffix), up(p.designation), safe(p.contact_number)])
            row += 1
        row += 1
    for i, w in enumerate([14, 12, 10, 18, 18, 12, 10, 25, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    safe_base = (file_name or 'form201_report').strip() or 'form201_report'
    return StreamingResponse(
        stream,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment;filename={safe_base}.xlsx"},
    )


@router.post('/form201-pdf')
def generate_form201_pdf(
    unit: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    prepared_by: str = Form(''),
    verified_by: str = Form(''),
    noted_by: str = Form(''),
    prepared_by_title: str = Form(''),
    verified_by_title: str = Form(''),
    noted_by_title: str = Form(''),
    file_name: str = Form('form201_report'),
    mask_birthdate: bool = Form(False),
    db: Session = Depends(get_db),
):
    """
    Generate a PDF Form 201 personnel master list grouped by sections.
    Columns (exact order): No., Badge No., Rank, Last Name, First Name, Middle Name, QLF,
    Date of Reassignment, Designation, Date of Designation, Mandatory Training,
    Specialized Course, Highest Eligibility, Contact Number, Birthdate, Religion
    """
    # Query personnel with filters
    query = db.query(models.Personnel)
    def up(v):
        return (v or '').strip().upper()

    def unit_key(raw_unit: str):
        u = up(raw_unit)
        if u in ('RHQ', 'RFU4A HQS', 'REGIONAL OFFICE', 'HEADQUARTERS'):
            return 'RHQ'
        if 'CAV' in u:
            return 'CAVITE'
        if 'LAG' in u:
            return 'LAGUNA'
        if 'BAT' in u:
            return 'BATANGAS'
        if 'RIZ' in u:
            return 'RIZAL'
        if 'QZN' in u or 'QUE' in u:
            return 'QUEZON'
        return u

    def unit_variants(raw_unit: Optional[str]):
        key = unit_key(raw_unit)
        variants = {
            'RHQ': ['RHQ', 'RFU4A HQS', 'REGIONAL OFFICE', 'HEADQUARTERS'],
            'CAVITE': ['CAVITE', 'CAVITE PFU', 'CAV PFU'],
            'LAGUNA': ['LAGUNA', 'LAGUNA PFU', 'LAG PFU'],
            'BATANGAS': ['BATANGAS', 'BATANGAS PFU', 'BATS PFU'],
            'RIZAL': ['RIZAL', 'RIZAL PFU'],
            'QUEZON': ['QUEZON', 'QUEZON PFU', 'QZN PFU'],
        }
        return variants.get(key, [key] if key else [])

    if unit and unit != 'All Units':
        unit_opts = unit_variants(unit)
        if unit_opts:
            query = query.filter(func.upper(models.Personnel.unit).in_(unit_opts))
    if status and status != 'All Status':
        query = query.filter(func.upper(models.Personnel.status) == up(status))

    all_personnel = query.all()

    # Group by section in required order while ensuring all records are represented
    rhq_sections = [
        'Office of the Regional Chief',
        'Admin and HRDD Section',
        'Intelligence Section',
        'Investigation Section',
        'Operation & PCR Section',
    ]
    non_rhq_sections = ['Cavite', 'Laguna', 'Batangas', 'Rizal', 'Quezon']
    ordered_sections = rhq_sections + non_rhq_sections
    grouped = {s: [] for s in ordered_sections}

    unit_display_map = {
        'CAVITE': 'Cavite',
        'LAGUNA': 'Laguna',
        'BATANGAS': 'Batangas',
        'RIZAL': 'Rizal',
        'QUEZON': 'Quezon',
    }

    for p in all_personnel:
        key = unit_key(getattr(p, 'unit', ''))
        if key == 'RHQ':
            section_name = (getattr(p, 'section', '') or '').strip()
            if section_name not in rhq_sections:
                section_name = 'Office of the Regional Chief'
        else:
            section_name = unit_display_map.get(key, 'Office of the Regional Chief')
        grouped.setdefault(section_name, []).append(p)

    # Build PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle('H', parent=styles['Heading2'], alignment=TA_CENTER, fontSize=12, leading=14)
    small_bold = ParagraphStyle('SB', parent=styles['Normal'], fontSize=8, leading=10)
    normal_small = ParagraphStyle('NS', parent=styles['Normal'], fontSize=7, leading=9)

    # Title
    elements.append(Paragraph('CRIMINAL INVESTIGATION AND DETECTION GROUP REGION 4A', header_style))
    elements.append(Paragraph('FORM 201 PERSONNEL MASTER LIST', ParagraphStyle('sub', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)))
    elements.append(Paragraph(f"(as of {datetime.utcnow().strftime('%B %d, %Y')})", ParagraphStyle('sub2', parent=styles['Normal'], alignment=TA_CENTER, fontSize=8)))
    elements.append(Spacer(1, 12))

    col_headers = ['No.', 'Badge No.', 'Rank', 'Last Name', 'First Name', 'Middle Name', 'QLF', 'Date of Reassignment', 'Designation', 'Date of Designation', 'Mandatory Training', 'Specialized Course', 'Highest Eligibility', 'Contact Number', 'Birthdate', 'Religion']

    # Column widths (points) - tuned to fit landscape letter with margins
    total_width = landscape(letter)[0] - 36  # left+right margins
    # distribute widths, favoring name columns
    col_widths = [28, 64, 48, 84, 84, 40, 44, 60, 64, 60, 80, 80, 64, 64, 56, 44]

    for section in ordered_sections:
        persons = grouped.get(section, [])
        if not persons:
            continue

        persons = sorted(
            persons,
            key=lambda p: (
                (getattr(p, 'rank', '') or '').upper(),
                (getattr(p, 'last_name', '') or '').upper(),
                (getattr(p, 'first_name', '') or '').upper(),
                (getattr(p, 'mi', '') or '').upper(),
            )
        )

        # Section header
        sect_para = Paragraph(f'<b>SECTION: {section}</b>', ParagraphStyle('sect', parent=styles['Normal'], fontSize=10))
        elements.append(sect_para)
        elements.append(Spacer(1, 6))

        # Table data
        data = [col_headers]
        for idx, p in enumerate(persons, start=1):
            # compile mandatory and specialized training titles
            mandatory_titles = [t.title for t in (p.trainings or []) if t.category == 'mandatory']
            specialized_titles = [t.title for t in (p.trainings or []) if t.category == 'specialized']

            def safe(val):
                if val is None or (isinstance(val, str) and not val.strip()):
                    return 'N/A'
                return val

            # format dates
            def fmt_date(val):
                if not val:
                    return 'N/A'
                try:
                    return val.strftime('%Y-%m-%d')
                except Exception:
                    return str(val)

            birth = '########' if mask_birthdate and p.birthdate else (fmt_date(p.birthdate) if p.birthdate else 'N/A')

            row = [
                idx,
                safe(p.badge_number),
                safe(p.rank),
                safe(p.last_name),
                safe(p.first_name),
                safe(p.mi),
                safe(p.qlf),
                fmt_date(p.date_of_reassignment) if getattr(p, 'date_of_reassignment', None) else 'N/A',
                safe(p.designation),
                fmt_date(p.date_of_designation) if getattr(p, 'date_of_designation', None) else 'N/A',
                '; '.join(mandatory_titles) if mandatory_titles else 'N/A',
                '; '.join(specialized_titles) if specialized_titles else 'N/A',
                safe(p.highest_eligibility),
                safe(p.contact_number),
                birth,
                safe(p.religion),
            ]
            data.append(row)

        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 8),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,1), (-1,-1), 7),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ])
        t.setStyle(style)
        elements.append(t)
        elements.append(Spacer(1, 12))

    # Footer signatories
    signatory_rows = [[
        Paragraph('<b>Prepared by:</b>', ParagraphStyle('sig_lbl', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)),
        Paragraph('<b>Verified Correct by:</b>', ParagraphStyle('sig_lbl', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)),
        Paragraph('<b>Noted by:</b>', ParagraphStyle('sig_lbl', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)),
    ], [
        Paragraph(f"<b>{(prepared_by or '').upper()}</b>", ParagraphStyle('sig_name', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)),
        Paragraph(f"<b>{(verified_by or '').upper()}</b>", ParagraphStyle('sig_name', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)),
        Paragraph(f"<b>{(noted_by or '').upper()}</b>", ParagraphStyle('sig_name', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER)),
    ], [
        Paragraph((prepared_by_title or '').upper(), ParagraphStyle('sig_title', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)),
        Paragraph((verified_by_title or '').upper(), ParagraphStyle('sig_title', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)),
        Paragraph((noted_by_title or '').upper(), ParagraphStyle('sig_title', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER)),
    ]]

    signatory_table = Table(signatory_rows, colWidths=[(total_width / 3)] * 3)
    signatory_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LINEABOVE', (0, 1), (-1, 1), 0.5, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(Spacer(1, 8))
    elements.append(signatory_table)

    doc.build(elements)
    buffer.seek(0)
    output_name = (file_name or 'form201_report').strip() or 'form201_report'
    return StreamingResponse(buffer, media_type='application/pdf', headers={"Content-Disposition": f"attachment;filename={output_name}.pdf"})
