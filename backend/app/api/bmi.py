from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from typing import Dict, List, Optional, Tuple
from sqlalchemy import text, extract, func
from sqlalchemy.orm import Session
from .. import models, schemas
from ..models import BMIHistory
from ..database import SessionLocal
from ..utils import (
    ensure_upload_folders,
    bmi_folder_name,
    safe_path_component,
    safe_resolve_upload_path,
    uploads_abs,
    uploads_rel,
)
import os
from datetime import datetime
import calendar
from io import BytesIO
from pathlib import Path
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

router = APIRouter()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def safe_filename(name: Optional[str], fallback: str = 'file') -> str:
    if not name:
        return fallback
    s = str(name).strip()
    # allow alphanumerics, dash, underscore, dot
    cleaned = ''.join(ch if (ch.isalnum() or ch in '-_.') else '_' for ch in s)
    cleaned = cleaned.strip('_.')
    return cleaned or fallback

# NOTE: Database migrations are executed on application startup
# (see `app.database.migrate_db()` called from `backend/main.py`).

PNP_BMI_AGE_TABLE = [
    {"group": "29 years old and below", "min_age": 0, "max_age": 29, "acceptable_min": 25.0, "acceptable_max": 29.9},
    {"group": "30-34 years old", "min_age": 30, "max_age": 34, "acceptable_min": 25.0, "acceptable_max": 29.9},
    {"group": "35-39 years old", "min_age": 35, "max_age": 39, "acceptable_min": 25.5, "acceptable_max": 29.9},
    {"group": "40 years old and above", "min_age": 40, "max_age": 200, "acceptable_min": 25.5, "acceptable_max": 29.9},
]

from .bmi_strict import draw_record_pdf_page
import logging

logger = logging.getLogger(__name__)

@router.post('/', response_model=schemas.BMISchema)
async def create_bmi(
    rank: str = Form(...),
    name: str = Form(...),
    first_name: Optional[str] = Form(None),
    last_name: Optional[str] = Form(None),
    mi: Optional[str] = Form(None),
    suffix: Optional[str] = Form(None),
    personnel_id: Optional[int] = Form(None),
    unit: str = Form(...),
    age: int = Form(...),
    sex: str = Form(...),
    height_cm: float = Form(...),
    weight_kg: float = Form(...),
    waist_cm: Optional[float] = Form(None),
    hip_cm: Optional[float] = Form(None),
    wrist_cm: Optional[float] = Form(None),
    date_taken: Optional[str] = Form(None),
    status: str = Form('Active'),
    status_custom: Optional[str] = Form(None),
    photo_front: UploadFile = File(...),
    photo_left: UploadFile = File(...),
    photo_right: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    ensure_upload_folders()
    # Normalize textual inputs to uppercase
    rank = rank.upper() if isinstance(rank, str) else rank
    name = name.upper() if isinstance(name, str) else name
    unit = unit.upper() if isinstance(unit, str) else unit
    sex = sex.upper() if isinstance(sex, str) else sex
    first_name = first_name.upper() if isinstance(first_name, str) else first_name
    last_name = last_name.upper() if isinstance(last_name, str) else last_name
    mi = mi.upper() if isinstance(mi, str) else mi
    suffix = suffix.upper() if isinstance(suffix, str) else suffix

    resolved_personnel_id = resolve_personnel_id(db, personnel_id, first_name, last_name, suffix, unit, mi)

    # Use explicit name parts when provided (more reliable than parsing full name)
    parts = name.strip().split() if name else []
    first_name_for_path = (first_name or '').strip() or (parts[0] if parts else '')
    last_name_for_path = (last_name or '').strip() or (parts[-1] if len(parts) > 1 else '')

    safe_unit_folder = safe_path_component(unit, 'UNIT').upper()
    safe_first = safe_path_component(first_name_for_path, 'FIRST').upper()
    safe_last = safe_path_component(last_name_for_path, 'LAST').upper()
    safe_mi = safe_path_component(mi, 'MI').upper() if mi and mi.strip() else ''
    folder_abs = uploads_abs('bmi', safe_unit_folder, bmi_folder_name(safe_first, safe_last, safe_mi))
    folder_rel = uploads_rel('bmi', safe_unit_folder, bmi_folder_name(safe_first, safe_last, safe_mi))
    os.makedirs(folder_abs, exist_ok=True)
    base_name = f"{safe_first}_{safe_mi}_{safe_last}" if safe_mi and safe_last else (f"{safe_first}_{safe_last}" if safe_last else safe_first)
    front_name = f"BMI_{base_name}_front.jpg"
    left_name = f"BMI_{base_name}_left.jpg"
    right_name = f"BMI_{base_name}_right.jpg"
    front_path_abs = os.path.join(str(folder_abs), front_name)
    left_path_abs = os.path.join(str(folder_abs), left_name)
    right_path_abs = os.path.join(str(folder_abs), right_name)

    bmi_value = compute_bmi(weight_kg, height_cm)
    pnp_classification = classify_pnp_bmi(bmi_value, age)
    parsed_date_taken = parse_date_taken(date_taken)
    
    # Determine the effective date (today if not provided)
    effective_date = parsed_date_taken or datetime.utcnow()

    # Mark any previously-latest records for this personnel/name as NOT latest
    db.query(models.BMIRecord).filter(
        models.BMIRecord.name.ilike(name),
        models.BMIRecord.is_latest == True,
    ).update({models.BMIRecord.is_latest: False}, synchronize_session=False)
    db.commit()

    record = models.BMIRecord(
        personnel_id=resolved_personnel_id,
        rank=rank, name=name, unit=unit, age=age, sex=sex,
        height_cm=height_cm, weight_kg=weight_kg, waist_cm=waist_cm,
        hip_cm=hip_cm, wrist_cm=wrist_cm, bmi=bmi_value, classification=pnp_classification,
        result=f"{bmi_value:.2f}",
        date_taken=parsed_date_taken or datetime.utcnow(),
        photo_front=f"{folder_rel}/{front_name}".replace("\\","/"),
        photo_left=f"{folder_rel}/{left_name}".replace("\\","/"),
        photo_right=f"{folder_rel}/{right_name}".replace("\\","/"),
        status=status,
        status_custom=status_custom,
        is_latest=True,  # Mark as latest
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    # If we linked to an existing Personnel and MI was provided, persist it if not already set
    if resolved_personnel_id and mi and (mi.strip()):
        try:
            person = db.query(models.Personnel).filter(models.Personnel.id == resolved_personnel_id).first()
            if person and (not getattr(person, 'mi', None) or not str(person.mi).strip()):
                person.mi = mi
                db.add(person)
                db.commit()
        except Exception:
            pass

    # Upsert BMI history snapshot for the month (used by the 14-month PDF timeline)
    if record.personnel_id:
        history = db.query(BMIHistory).filter(
            BMIHistory.personnel_id == record.personnel_id,
            BMIHistory.month == effective_date.month,
            BMIHistory.year == effective_date.year,
        ).first()
        if history:
            history.weight = float(weight_kg)
            history.bmi_result = float(bmi_value)
            history.bmi_classification = pnp_classification
            history.waist = waist_cm
            history.hip = hip_cm
            history.wrist = wrist_cm
            history.created_at = effective_date
        else:
            history = BMIHistory(
                personnel_id=record.personnel_id,
                month=effective_date.month,
                year=effective_date.year,
                weight=float(weight_kg),
                bmi_result=float(bmi_value),
                bmi_classification=pnp_classification,
                waist=waist_cm,
                hip=hip_cm,
                wrist=wrist_cm,
                created_at=effective_date,
            )
            db.add(history)
        db.commit()

    # save photos after record exists
    with open(front_path_abs,'wb') as f:
        f.write(await photo_front.read())
    with open(left_path_abs,'wb') as f:
        f.write(await photo_left.read())
    with open(right_path_abs,'wb') as f:
        f.write(await photo_right.read())

    return record


@router.post('/report')
def bmi_report(
    month: Optional[int] = Form(None),
    year: Optional[int] = Form(None),
    unit: Optional[str] = Form(None),
    prepared_by: str = Form(''),
    verified_by: str = Form(''),
    noted_by: str = Form(''),
    status: Optional[str] = Form(None),
    report_type: str = Form('pdf'),
    file_name: str = Form('bmi_report'),
    db: Session = Depends(get_db),
):
    # Debug: print incoming payload for traceability
    print(f"BMI report request payload: month={month}, year={year}, unit={unit}, report_type={report_type}, file_name={file_name}, prepared_by={prepared_by}, verified_by={verified_by}, noted_by={noted_by}, status={status}")

    # Top-level try/except to ensure full traceback logging on any failure
    try:
        # Store report month/year for PDF generation
        report_month = month
        report_year = year

        if not month or not year:
            raise HTTPException(status_code=400, detail='Month and Year are required for BMI reports.')

        # fetch records matching filters
        q = db.query(models.BMIRecord)
        if unit and unit != 'All Units':
            unit_opts = unit_variants(unit)
            if unit_opts:
                q = q.filter(func.upper(models.BMIRecord.unit).in_(unit_opts))

        # filter by date_taken month/year
        q = q.filter(extract('month', models.BMIRecord.date_taken) == month, extract('year', models.BMIRecord.date_taken) == year)
        records = q.order_by(models.BMIRecord.date_taken.desc(), models.BMIRecord.id.desc()).all()

        # Filter BMI records based on status parameters
        filtered_records = []
        for rec in records:
            rec_status = (getattr(rec, 'status', None) or '').strip()
            if status:
                if rec_status and rec_status.upper() == status.upper():
                    filtered_records.append(rec)
                else:
                    if not rec_status and status.upper() == 'ACTIVE':
                        filtered_records.append(rec)
            else:
                if not rec_status or rec_status.upper() == 'ACTIVE':
                    filtered_records.append(rec)

        records = filtered_records
        safe_base_name = safe_filename(file_name, 'bmi_report')

        # De-dupe: keep only the most recent entry per personnel/name
        seen = set()
        deduped = []
        for rec in records:
            key = rec.personnel_id or ((rec.name or '').strip().upper() or rec.id)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(rec)

        records = deduped

        if not records:
            raise HTTPException(status_code=404, detail='No BMI records found for the selected Month/Year/Unit filters.')

        if report_type.lower() == 'excel':
            ids = [r.id for r in records]
            return generate_excel(
                records=ids,
                prepared_by=prepared_by,
                verified_by=verified_by,
                noted_by=noted_by,
                file_name=safe_base_name,
                report_month=month,
                report_year=year,
                report_unit=(unit or 'All Units'),
                db=db,
            )

        # For PDF: create combined multi-page PDF
        buffer = BytesIO()
        try:
            # Use exact A4 landscape dimensions: 842 x 595 pts
            c = canvas.Canvas(buffer, pagesize=(842, 595))
            for rec in records:
                draw_record_pdf_page(c, rec, db=db, prepared_by=prepared_by, verified_by=verified_by, noted_by=noted_by,
                                    report_month=report_month, report_year=report_year)
                c.showPage()

            c.save()
            buffer.seek(0)
            return StreamingResponse(
                buffer,
                media_type='application/pdf',
                headers={"Content-Disposition": f"attachment;filename={safe_base_name}.pdf"},
            )
        except Exception as e:
            import traceback
            print("BMI REPORT ERROR during PDF generation:", str(e))
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=str(e))
        finally:
            # Do not close the buffer here; StreamingResponse will consume it.
            pass

    except Exception as e:
        import traceback
        print('BMI REPORT ERROR:', str(e))
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get('/{record_id}/pdf')
def generate_bmi_pdf(record_id: int, file_name: Optional[str] = None, report_month: Optional[int] = None, report_year: Optional[int] = None, db: Session = Depends(get_db)):
    rec = db.query(models.BMIRecord).filter(models.BMIRecord.id==record_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail='BMI record not found')

    buffer = BytesIO()
    # Use exact A4 landscape dimensions: 842 x 595 pts
    c = canvas.Canvas(buffer, pagesize=(842, 595))
    draw_record_pdf_page(c, rec, db=db, report_month=report_month, report_year=report_year)
    c.showPage()
    c.save()
    buffer.seek(0)
    output_name = safe_filename(file_name or f"bmi_{record_id}", f"bmi_{record_id}")
    return StreamingResponse(
        buffer,
        media_type='application/pdf',
        headers={"Content-Disposition": f"attachment;filename={output_name}.pdf"},
    )


@router.post('/excel')
def generate_excel(
    records: List[int] = Form(...),
    prepared_by: str = Form(''),
    verified_by: str = Form(''),
    noted_by: str = Form(''),
    file_name: str = Form('bmi_report'),
    report_month: Optional[int] = Form(None),
    report_year: Optional[int] = Form(None),
    report_unit: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    from collections import defaultdict
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.page import PageMargins

    def _sheet_print_setup(ws, header_row: int) -> None:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.3, footer=0.3)
        ws.print_options.horizontalCentered = True
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = ws[f"A{header_row + 1}"]
        ws.print_title_rows = f"{header_row}:{header_row}"

    def _auto_fit_columns(ws, max_col: int, min_width: int = 10, max_width: int = 40) -> None:
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for c in cell:
                    if c.value is None:
                        continue
                    max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))

    def _apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int, border: Border) -> None:
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = border

    def _write_report_header(ws, title: str, col_count: int, period_text: str, unit_text: str) -> None:
        last_col = get_column_letter(col_count)
        ws.merge_cells(f"A1:{last_col}1")
        ws['A1'].value = title
        ws['A1'].font = Font(name='Calibri', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22

        ws.merge_cells(f"A2:{last_col}2")
        ws['A2'].value = period_text
        ws['A2'].font = Font(name='Calibri', size=11, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(f"A3:{last_col}3")
        ws['A3'].value = unit_text
        ws['A3'].font = Font(name='Calibri', size=11, bold=True)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

        ws.append([])  # spacer row

    # Fetch all records
    recs = []
    for rid in records:
        rec = db.query(models.BMIRecord).filter(models.BMIRecord.id == rid).first()
        if rec:
            recs.append(rec)

    wb = Workbook()

    period_str = ""
    if report_month and report_year:
        try:
            period_str = f"{calendar.month_name[int(report_month)]} {int(report_year)}"
        except Exception:
            period_str = f"{report_month} {report_year}"
    unit_scope_str = (report_unit or 'All Units')

    # Shared styles
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')
    header_font = Font(name='Calibri', size=11, bold=True)
    body_font = Font(name='Calibri', size=11)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # RECAP Sheet
    recap_ws = wb.active
    recap_ws.title = 'RECAP'
    recap_headers = ['Rank', 'Underweight', 'Normal', 'Overweight', 'Obese 1', 'Obese 2', 'Obese 3']
    _write_report_header(
        recap_ws,
        title='BMI RECAP REPORT (BY RANK)',
        col_count=len(recap_headers),
        period_text=(f"Report Period: {period_str}" if period_str else 'Report Period: -'),
        unit_text=f"Unit Scope: {unit_scope_str}",
    )

    header_row = recap_ws.max_row + 1
    for col_idx, h in enumerate(recap_headers, 1):
        cell = recap_ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # Compute rank counts for recap
    rank_counts = defaultdict(lambda: defaultdict(int))
    total_counts = defaultdict(int)

    for rec in recs:
        bmi_value = compute_bmi(rec.weight_kg or 0, rec.height_cm or 0)
        pnp_classification = classify_pnp_bmi(bmi_value, rec.age or 0)

        # Map classifications to RECAP columns
        cls_map = {
            'Severely Underweight': 'Underweight',
            'Underweight': 'Underweight',
            'Normal': 'Normal',
            'Acceptable BMI': 'Normal',
            'Overweight': 'Overweight',
            'Obese Class 1': 'Obese Class 1',
            'Obese Class 2': 'Obese Class 2',
            'Obese Class 3': 'Obese Class 3'
        }
        cls = cls_map.get(pnp_classification, 'Normal')  # Default to Normal if unknown

        rank = rec.rank or 'Unknown'
        rank_counts[rank][cls] += 1
        total_counts[cls] += 1

    # Add rows to recap (sorted by standard rank order, including zero count ranks)
    standard_ranks = [
        'PGEN', 'PLTGEN', 'PMGEN', 'PBGEN', 'PCOL', 'PLTCOL', 'PMAJ', 'PCPT', 'PLT',
        'PEMS', 'PCMS', 'PSMS', 'PMSG', 'PSSG', 'PCPL', 'PAT', 'NUP'
    ]
    other_ranks = sorted(set(rank_counts.keys()) - set(standard_ranks))
    recap_rows_start = header_row + 1
    current_row = recap_rows_start
    for rank in standard_ranks + other_ranks:
        recap_ws.cell(row=current_row, column=1, value=rank).alignment = align_left
        recap_ws.cell(row=current_row, column=1).font = body_font
        for idx, cls in enumerate(['Underweight', 'Normal', 'Overweight', 'Obese Class 1', 'Obese Class 2', 'Obese Class 3'], start=2):
            c = recap_ws.cell(row=current_row, column=idx, value=rank_counts[rank][cls])
            c.font = body_font
            c.alignment = align_center
        current_row += 1

    # Add TOTAL row
    total_row_idx = current_row
    recap_ws.cell(row=total_row_idx, column=1, value='TOTAL').font = Font(name='Calibri', size=11, bold=True)
    recap_ws.cell(row=total_row_idx, column=1).alignment = align_left
    for idx, cls in enumerate(['Underweight', 'Normal', 'Overweight', 'Obese Class 1', 'Obese Class 2', 'Obese Class 3'], start=2):
        c = recap_ws.cell(row=total_row_idx, column=idx, value=total_counts[cls])
        c.font = Font(name='Calibri', size=11, bold=True)
        c.alignment = align_center

    # Borders + table style for recap
    last_col_letter = get_column_letter(len(recap_headers))
    last_row = total_row_idx
    _apply_border(recap_ws, header_row, last_row, 1, len(recap_headers), thin_border)
    recap_table_ref = f"A{header_row}:{last_col_letter}{last_row}"
    recap_table = Table(displayName='Tbl_RECAP', ref=recap_table_ref)
    recap_table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    recap_ws.add_table(recap_table)
    _sheet_print_setup(recap_ws, header_row)
    _auto_fit_columns(recap_ws, max_col=len(recap_headers), min_width=10, max_width=28)

    # Unit sheets
    unit_sheets = [
        ('RHQ', 'RHQ'),
        ('CAVITE', 'Cavite'),
        ('LAGUNA', 'Laguna'),
        ('BATANGAS', 'Batangas'),
        ('RIZAL', 'Rizal'),
        ('QUEZON', 'Quezon'),
    ]

    selected_unit_code = None
    if report_unit and str(report_unit).strip() and str(report_unit).strip().upper() != 'ALL UNITS':
        selected_unit_code = unit_key(str(report_unit))

    if selected_unit_code:
        unit_sheets_to_make = [u for u in unit_sheets if u[0] == selected_unit_code]
        if not unit_sheets_to_make:
            unit_sheets_to_make = unit_sheets
    else:
        unit_sheets_to_make = unit_sheets

    for unit_code, sheet_name in unit_sheets_to_make:
        unit_recs = [r for r in recs if unit_key(getattr(r, 'unit', None)) == unit_code]
        # Always create the sheet, even if empty
        ws = wb.create_sheet(sheet_name)
        unit_headers = ['No.', 'Rank/Name', 'Height (cm)', 'Weight (kg)', 'Age', 'BMI Result', 'BMI Classification', 'Weight to Lose (kg)']

        _write_report_header(
            ws,
            title=f"BMI MONITORING REPORT - {sheet_name.upper()}",
            col_count=len(unit_headers),
            period_text=(f"Report Period: {period_str}" if period_str else 'Report Period: -'),
            unit_text=f"Unit Scope: {sheet_name}",
        )

        header_row = ws.max_row + 1
        for col_idx, h in enumerate(unit_headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        # Sort for a stable, official-looking output
        unit_recs_sorted = sorted(unit_recs, key=lambda r: ((r.rank or ''), (r.name or '')))

        data_start_row = header_row + 1
        current_row = data_start_row
        for i, rec in enumerate(unit_recs_sorted, 1):
            bmi_value = compute_bmi(rec.weight_kg or 0, rec.height_cm or 0)
            pnp_classification = classify_pnp_bmi(bmi_value, rec.age or 0)
            metrics = compute_weight_metrics(rec.height_cm or 0, rec.weight_kg or 0, rec.age or 0)
            weight_to_lose = float(metrics['weight_to_lose']) if metrics.get('weight_to_lose') and metrics['weight_to_lose'] > 0 else None

            row_values = [
                i,
                f"{rec.rank or ''} {rec.name or ''}".strip(),
                float(rec.height_cm) if rec.height_cm is not None else None,
                float(rec.weight_kg) if rec.weight_kg is not None else None,
                int(rec.age) if rec.age is not None else None,
                float(bmi_value),
                pnp_classification,
                weight_to_lose,
            ]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = body_font
                if col_idx == 2 or col_idx == 7:
                    cell.alignment = align_left
                elif col_idx in (1, 5):
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right

                # Number formats
                if col_idx == 3:
                    cell.number_format = '0.0'
                if col_idx == 4:
                    cell.number_format = '0.0'
                if col_idx == 5:
                    cell.number_format = '0'
                if col_idx == 6:
                    cell.number_format = '0.00'
                if col_idx == 8:
                    cell.number_format = '0.0'

            current_row += 1

        # Apply borders + table style (only over header+data)
        last_data_row = current_row - 1
        if last_data_row < header_row:
            last_data_row = header_row
        _apply_border(ws, header_row, last_data_row, 1, len(unit_headers), thin_border)

        table_name = f"Tbl_{sheet_name.upper()}"
        table_ref = f"A{header_row}:{get_column_letter(len(unit_headers))}{last_data_row}"
        unit_table = Table(displayName=table_name, ref=table_ref)
        unit_table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium9',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(unit_table)

        # Signatories at bottom (Excel only)
        ws.append([])
        ws.append([])
        sig_start = ws.max_row + 1
        ws.cell(row=sig_start, column=1, value='Prepared by:').font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=sig_start, column=2, value=prepared_by).font = body_font
        ws.cell(row=sig_start + 1, column=1, value='Verified Correct by:').font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=sig_start + 1, column=2, value=verified_by).font = body_font
        ws.cell(row=sig_start + 2, column=1, value='Noted by:').font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=sig_start + 2, column=2, value=noted_by).font = body_font

        _sheet_print_setup(ws, header_row)
        _auto_fit_columns(ws, max_col=len(unit_headers), min_width=10, max_width=38)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    safe_base_name = safe_filename(file_name, 'bmi_report')
    return StreamingResponse(
        stream,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment;filename={safe_base_name}.xlsx"},
    )


# ==================== BMI LIST AND HISTORY ENDPOINTS ====================

@router.get('/', response_model=List[schemas.BMISchema])
def list_bmi(
    personnel_id: Optional[int] = None,
    unit: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    exact_date: Optional[str] = None,
    search: Optional[str] = None,
    latest_only: bool = False,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    List BMI records with optional filtering.
    
    Query params:
    - personnel_id: Filter by personnel ID
    - unit: Filter by unit (RHQ, Cavite, etc.)
    - month: Filter by month (1-12)
    - year: Filter by year
    - exact_date: Filter by exact date (YYYY-MM-DD)
    - search: Search by name or rank
    - latest_only: If true, return only records where is_latest=True (one per personnel)
    """
    from sqlalchemy import extract, func, or_
    
    q = db.query(models.BMIRecord)
    
    # Apply filters
    if personnel_id:
        q = q.filter(models.BMIRecord.personnel_id == personnel_id)
    
    if unit and unit != 'All Units':
        unit_opts = unit_variants(unit)
        if unit_opts:
            q = q.filter(func.upper(models.BMIRecord.unit).in_(unit_opts))
    
    if month and year:
        q = q.filter(
            extract('month', models.BMIRecord.date_taken) == month,
            extract('year', models.BMIRecord.date_taken) == year
        )
    
    if exact_date:
        parsed_date = parse_date_taken(exact_date)
        if parsed_date:
            q = q.filter(
                extract('year', models.BMIRecord.date_taken) == parsed_date.year,
                extract('month', models.BMIRecord.date_taken) == parsed_date.month,
                extract('day', models.BMIRecord.date_taken) == parsed_date.day
            )
    
    if search:
        search_term = f"%{search}%"
        q = q.filter(
            or_(
                models.BMIRecord.name.ilike(search_term),
                models.BMIRecord.rank.ilike(search_term)
            )
        )
    
    if latest_only:
        # NEW: Use is_latest flag for filtering - reliable one-per-personnel approach
        q = q.filter(models.BMIRecord.is_latest == True)

    # Apply status filtering: if provided, use it; otherwise default to Active or unspecified
    if status:
        q = q.filter(func.upper(models.BMIRecord.status) == status.upper())
    else:
        from sqlalchemy import or_
        q = q.filter(or_(models.BMIRecord.status == None, func.upper(models.BMIRecord.status) == 'ACTIVE'))
    
    return q.order_by(models.BMIRecord.date_taken.desc()).all()


@router.get('/history/{identifier}', response_model=schemas.BMIHistorySchema)
def get_bmi_history(
    identifier: str,
    db: Session = Depends(get_db)
):
    """
    Get all BMI history records for a specific personnel.
    
    The identifier can be:
    - A BMI record ID (numeric): Finds that record and returns all records with the same name
    - A personnel ID from Personnel table (numeric)
    - A name string: Finds all BMI records matching that name
    
    Returns the personnel info, latest record, and full history sorted newest first.
    """
    records = []
    personnel = None
    personnel_name = None
    personnel_rank = None
    personnel_unit = None
    
    # Check if it's a numeric ID or a name
    is_numeric = identifier.isdigit()
    
    if is_numeric:
        record_id = int(identifier)
        
        # First, try to find by BMI record ID and get its name
        bmi_record = db.query(models.BMIRecord).filter(models.BMIRecord.id == record_id).first()
        
        if bmi_record:
            # Get the personnel name from this BMI record
            personnel_name = bmi_record.name
            personnel_rank = bmi_record.rank
            personnel_unit = bmi_record.unit
            
            # Find all BMI records with the same name (case-insensitive)
            records = db.query(models.BMIRecord).filter(
                models.BMIRecord.name.ilike(bmi_record.name)
            ).order_by(models.BMIRecord.date_taken.desc()).all()
            
            # Also try to find Personnel by this name
            personnel = db.query(models.Personnel).filter(
                (models.Personnel.first_name + ' ' + models.Personnel.last_name).ilike(f"%{personnel_name}%") |
                (models.Personnel.last_name + ' ' + models.Personnel.first_name).ilike(f"%{personnel_name}%")
            ).first()
        else:
            # ID not found as BMI record, try it as a Personnel ID
            personnel = db.query(models.Personnel).filter(models.Personnel.id == record_id).first()
            if personnel:
                personnel_name = f"{personnel.first_name} {personnel.last_name}"
                personnel_rank = personnel.rank
                personnel_unit = personnel.unit
                # Get all BMI records for this personnel by personnel_id
                records = db.query(models.BMIRecord).filter(
                    models.BMIRecord.personnel_id == personnel.id
                ).order_by(models.BMIRecord.date_taken.desc()).all()
                
                # If no records by personnel_id, try by name
                if not records:
                    records = db.query(models.BMIRecord).filter(
                        models.BMIRecord.name.ilike(f"%{personnel_name}%")
                    ).order_by(models.BMIRecord.date_taken.desc()).all()
    else:
        # It's a name - find all matching BMI records
        personnel_name = identifier
        records = db.query(models.BMIRecord).filter(
            models.BMIRecord.name.ilike(f"%{identifier}%")
        ).order_by(models.BMIRecord.date_taken.desc()).all()
        
        if records:
            # Get info from the first (latest) record
            latest = records[0]
            personnel_rank = latest.rank
            personnel_unit = latest.unit
            
            # Try to find Personnel by name
            personnel = db.query(models.Personnel).filter(
                (models.Personnel.first_name + ' ' + models.Personnel.last_name).ilike(f"%{latest.name}%") |
                (models.Personnel.last_name + ' ' + models.Personnel.first_name).ilike(f"%{latest.name}%")
            ).first()
    
    # Get latest record
    latest = records[0] if records else None
    
    return {
        "personnel_id": personnel.id if personnel else (latest.personnel_id if latest else None),
        "personnel_name": personnel_name if personnel_name else (latest.name if latest else None),
        "personnel_rank": personnel_rank if personnel_rank else (latest.rank if latest else None),
        "personnel_unit": personnel_unit if personnel_unit else (latest.unit if latest else None),
        "latest_bmi": latest,
        "history": records
    }


@router.get('/history/by-name/{name}', response_model=schemas.BMIHistorySchema)
def get_bmi_history_by_name(
    name: str,
    db: Session = Depends(get_db)
):
    """
    Get BMI history by personnel name (for cases where personnel_id is not set).
    """
    # Get all BMI records matching this name
    records = db.query(models.BMIRecord).filter(
        models.BMIRecord.name.ilike(f"%{name}%")
    ).order_by(models.BMIRecord.date_taken.desc()).all()
    
    latest = records[0] if records else None
    
    return {
        "personnel_id": latest.personnel_id if latest else None,
        "personnel_name": name,
        "personnel_rank": latest.rank if latest else None,
        "personnel_unit": latest.unit if latest else None,
        "latest_bmi": latest,
        "history": records
    }


@router.get('/history/{personnel_id}/by-date', response_model=schemas.BMISchema)
def get_bmi_by_date(
    personnel_id: int,
    date: str,  # YYYY-MM-DD format
    db: Session = Depends(get_db)
):
    """
    Get the BMI record for a specific personnel on a specific date.
    If multiple records exist on that day, returns the most recent one.
    """
    from sqlalchemy import extract
    
    parsed_date = parse_date_taken(date)
    if not parsed_date:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Get the personnel info
    personnel = db.query(models.Personnel).filter(models.Personnel.id == personnel_id).first()
    
    # Try to find records for this personnel on the specific date
    records = db.query(models.BMIRecord).filter(
        models.BMIRecord.personnel_id == personnel_id,
        extract('year', models.BMIRecord.date_taken) == parsed_date.year,
        extract('month', models.BMIRecord.date_taken) == parsed_date.month,
            extract('day', models.BMIRecord.date_taken) == parsed_date.day
        ).order_by(models.BMIRecord.date_taken.desc()).all()
    
    # If no records linked by personnel_id, try by name matching
    if not records and personnel:
        name_pattern = f"%{personnel.first_name}%{personnel.last_name}%"
        records = db.query(models.BMIRecord).filter(
            models.BMIRecord.name.ilike(name_pattern),
            extract('year', models.BMIRecord.date_taken) == parsed_date.year,
            extract('month', models.BMIRecord.date_taken) == parsed_date.month,
            extract('day', models.BMIRecord.date_taken) == parsed_date.day
        ).order_by(models.BMIRecord.date_taken.desc()).all()
    
    if not records:
        raise HTTPException(
            status_code=404,
            detail=f"No BMI record found for the selected date {date}"
        )
    
    # Return the most recent one
    return records[0]


@router.get('/latest/{personnel_id}', response_model=schemas.BMISchema)
def get_latest_bmi(
    personnel_id: int,
    db: Session = Depends(get_db)
):
    """
    Get the latest BMI record for a specific personnel.
    """
    # Get the personnel info
    personnel = db.query(models.Personnel).filter(models.Personnel.id == personnel_id).first()
    
    # Try to find the latest record for this personnel
    record = db.query(models.BMIRecord).filter(
        models.BMIRecord.personnel_id == personnel_id
    ).order_by(models.BMIRecord.date_taken.desc()).first()
    
    # If no records linked by personnel_id, try by name matching
    if not record and personnel:
        name_pattern = f"%{personnel.first_name}%{personnel.last_name}%"
        record = db.query(models.BMIRecord).filter(
            models.BMIRecord.name.ilike(name_pattern)
        ).order_by(models.BMIRecord.date_taken.desc()).first()
    
    if not record:
        raise HTTPException(
            status_code=404,
            detail="No BMI records found for this personnel"
        )
    
    return record


@router.get('/timeline/{personnel_id}', response_model=List[schemas.BMITimelineSchema])
def get_bmi_timeline(
    personnel_id: int,
    db: Session = Depends(get_db)
):
    """
    Get simplified timeline data for BMI charts/cards.
    Returns date, BMI value, classification, and weight for each record.
    """
    # Get the personnel info
    personnel = db.query(models.Personnel).filter(models.Personnel.id == personnel_id).first()

    # Get all BMI records for timeline
    records = db.query(models.BMIRecord).filter(
        models.BMIRecord.personnel_id == personnel_id
    ).order_by(models.BMIRecord.date_taken.asc()).all()

    # If no records linked by personnel_id, try by name matching
    if not records and personnel:
        name_pattern = f"%{personnel.first_name}%{personnel.last_name}%"
        records = db.query(models.BMIRecord).filter(
            models.BMIRecord.name.ilike(name_pattern)
        ).order_by(models.BMIRecord.date_taken.asc()).all()

    timeline = []
    for rec in records:
        timeline.append({
            "date_taken": rec.date_taken,
            "bmi": rec.bmi,
            "classification": rec.classification,
            "weight_kg": rec.weight_kg
        })

    return timeline


@router.get('/history-data/{personnel_id}')
def get_bmi_history_data(personnel_id: int, report_month: int = None, report_year: int = None, db: Session = Depends(get_db)):
    """
    Get BMI history data for PDF reports with dynamic 14-month timeline.
    """
    if not report_month or not report_year:
        # Default to current month/year if not provided
        now = datetime.utcnow()
        report_month = report_month or now.month
        report_year = report_year or now.year

    # Get personnel info
    personnel = db.query(models.Personnel).filter(models.Personnel.id == personnel_id).first()
    if not personnel:
        raise HTTPException(status_code=404, detail="Personnel not found")

    # Get BMI history for this personnel
    history_records = db.query(BMIHistory).filter(
        BMIHistory.personnel_id == personnel_id
    ).order_by(BMIHistory.year.asc(), BMIHistory.month.asc()).all()

    # If bmi_history is empty (older deployments / legacy data), fall back to bmi_records.
    record_month_map: Dict[Tuple[int, int], models.BMIRecord] = {}
    if not history_records:
        legacy_records = db.query(models.BMIRecord).filter(
            models.BMIRecord.personnel_id == personnel_id
        ).order_by(models.BMIRecord.date_taken.desc(), models.BMIRecord.id.desc()).all()
        for r in legacy_records:
            if not r.date_taken:
                continue
            key = (r.date_taken.year, r.date_taken.month)
            if key not in record_month_map:
                record_month_map[key] = r

    # Build 14-month timeline with report month as the rightmost column
    months_data = []
    for i in range(13, -1, -1):  # 14 months backward from report month
        target_month = report_month - i
        target_year = report_year

        # Adjust for year boundaries
        while target_month <= 0:
            target_month += 12
            target_year -= 1
        while target_month > 12:
            target_month -= 12
            target_year += 1

        # Find matching history record
        matching_record = None
        for rec in history_records:
            if rec.month == target_month and rec.year == target_year:
                matching_record = rec
                break

        legacy_record = record_month_map.get((target_year, target_month)) if record_month_map else None

        months_data.append({
            "month": target_month,
            "year": target_year,
            "month_name": calendar.month_abbr[target_month],
            "weight": matching_record.weight if matching_record else (legacy_record.weight_kg if legacy_record else None),
            "bmi": matching_record.bmi_result if matching_record else (legacy_record.bmi if legacy_record else None),
            "classification": matching_record.bmi_classification if matching_record else (legacy_record.classification if legacy_record else None)
        })

    return {
        "personnel": {
            "id": personnel.id,
            "name": f"{personnel.first_name} {personnel.last_name}",
            "rank": personnel.rank,
            "unit": personnel.unit
        },
        "report_period": {
            "month": report_month,
            "year": report_year,
            "month_name": calendar.month_name[report_month]
        },
        "monthly_data": months_data
    }


@router.get('/personnel-list')
def get_personnel_with_bmi(db: Session = Depends(get_db)):
    """
    Get list of all personnel with their latest BMI status.
    Useful for dropdown selection in the BMI history view.
    """
    from sqlalchemy import func
    
    # Get all personnel
    all_personnel = db.query(models.Personnel).all()
    
    result = []
    for p in all_personnel:
        # Get latest BMI for this personnel
        latest_bmi = db.query(models.BMIRecord).filter(
            models.BMIRecord.personnel_id == p.id
        ).order_by(models.BMIRecord.date_taken.desc()).first()
        
        # Also check by name matching
        if not latest_bmi:
            name_pattern = f"%{p.first_name}%{p.last_name}%"
            latest_bmi = db.query(models.BMIRecord).filter(
                models.BMIRecord.name.ilike(name_pattern)
            ).order_by(models.BMIRecord.date_taken.desc()).first()
        
        # Count total BMI records for this personnel
        total_records = db.query(models.BMIRecord).filter(
            models.BMIRecord.personnel_id == p.id
        ).count()
        
        if not total_records:
            name_pattern = f"%{p.first_name}%{p.last_name}%"
            total_records = db.query(models.BMIRecord).filter(
                models.BMIRecord.name.ilike(name_pattern)
            ).count()
        
        result.append({
            "id": p.id,
            "name": f"{p.first_name} {p.last_name}",
            "rank": p.rank,
            "unit": p.unit,
            "status": p.status,
            "latest_bmi": latest_bmi.bmi if latest_bmi else None,
            "latest_classification": latest_bmi.classification if latest_bmi else None,
            "latest_date": latest_bmi.date_taken if latest_bmi else None,
            "total_records": total_records
        })
    
    return result


@router.get('/distinct-personnel')
def get_distinct_personnel_bmi(db: Session = Depends(get_db)):
    """
    Get list of all distinct personnel who have BMI records.
    Returns records where is_latest=True (one per personnel).
    """
    from sqlalchemy import func
    
    # NEW: Use is_latest flag - reliable approach
    # Get all records marked as latest
    latest_records = db.query(models.BMIRecord).filter(
        models.BMIRecord.is_latest == True
    ).all()
    
    result = []
    for rec in latest_records:
        # Count total BMI records for this personnel (by name)
        total_records = db.query(models.BMIRecord).filter(
            models.BMIRecord.name.ilike(rec.name)
        ).count()
        
        result.append({
            "id": rec.id,
            "name": (getattr(rec, 'display_name', None) or rec.name),
            "rank": rec.rank,
            "unit": rec.unit,
            "personnel_id": rec.personnel_id,
            "latest_bmi": rec.bmi,
            "latest_classification": rec.classification,
            "latest_date": rec.date_taken,
            "total_records": total_records,
            "is_latest": rec.is_latest
        })
    
    return sorted(result, key=lambda x: (x['name'] or '').lower())


@router.get('/counts')
def bmi_counts(db: Session = Depends(get_db)):
    """
    Get BMI record counts per unit for the current month and total counts.
    """
    from sqlalchemy import extract, func
    from datetime import datetime
    
    now = datetime.utcnow()
    current_month = now.month
    current_year = now.year
    
    unit_map = {
        'RHQ': 'RHQ',
        'Cavite': 'CAVITE',
        'Laguna': 'LAGUNA',
        'Batangas': 'BATANGAS',
        'Rizal': 'RIZAL',
        'Quezon': 'QUEZON',
    }
    result = {}
    
    for display_name, db_unit in unit_map.items():
        # Count records for this unit in current month
        monthly_count = db.query(models.BMIRecord).filter(
            func.upper(models.BMIRecord.unit) == db_unit,
            extract('month', models.BMIRecord.date_taken) == current_month,
            extract('year', models.BMIRecord.date_taken) == current_year
        ).count()
        
        # Count total records for this unit
        total_count = db.query(models.BMIRecord).filter(
            func.upper(models.BMIRecord.unit) == db_unit
        ).count()
        
        result[display_name] = {
            'monthly': monthly_count,
            'total': total_count
        }
    
    # Get total monthly count across all units
    total_monthly = db.query(models.BMIRecord).filter(
        extract('month', models.BMIRecord.date_taken) == current_month,
        extract('year', models.BMIRecord.date_taken) == current_year
    ).count()
    
    result['total_monthly'] = total_monthly
    result['total'] = sum(v.get('total', 0) for v in result.values() if isinstance(v, dict))
    result['current_month'] = now.strftime('%B %Y')
    
    return result


# ==================== BMI UPDATE ENDPOINT (Version-Preserving) ====================

@router.put('/{record_id}', response_model=schemas.BMISchema)
async def update_bmi(
    record_id: int,
    rank: str = Form(...),
    name: str = Form(...),
    first_name: Optional[str] = Form(None),
    last_name: Optional[str] = Form(None),
    mi: Optional[str] = Form(None),
    suffix: Optional[str] = Form(None),
    personnel_id: Optional[int] = Form(None),
    unit: str = Form(...),
    age: int = Form(...),
    sex: str = Form(...),
    height_cm: float = Form(...),
    weight_kg: float = Form(...),
    waist_cm: Optional[float] = Form(None),
    hip_cm: Optional[float] = Form(None),
    wrist_cm: Optional[float] = Form(None),
    date_taken: Optional[str] = Form(None),
    status: str = Form('Active'),
    status_custom: Optional[str] = Form(None),
    photo_front: Optional[UploadFile] = File(None),
    photo_left: Optional[UploadFile] = File(None),
    photo_right: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    """
    Update an existing BMI record - VERSION PRESERVING.
    
    IMPORTANT: This does NOT overwrite the old record. Instead:
    1. The OLD record is kept unchanged as a historical snapshot (is_latest = False)
    2. A NEW BMI record is created with the updated values (is_latest = True)
    3. The new record becomes the latest/current version
    
    This preserves the complete history while allowing corrections/updates.
    """
    # Find the existing (old) record to verify it exists
    old_record = db.query(models.BMIRecord).filter(models.BMIRecord.id == record_id).first()
    if not old_record:
        raise HTTPException(status_code=404, detail="BMI record not found")
    
    # Preserve the old record's important fields before we do anything
    preserved_personnel_id = old_record.personnel_id
    preserved_old_front = old_record.photo_front
    preserved_old_left = old_record.photo_left
    preserved_old_right = old_record.photo_right
    
    # normalize incoming text fields for consistent matching
    rank = rank.upper() if isinstance(rank, str) else rank
    name = name.upper() if isinstance(name, str) else name
    unit = unit.upper() if isinstance(unit, str) else unit
    sex = sex.upper() if isinstance(sex, str) else sex
    first_name = first_name.upper() if isinstance(first_name, str) else first_name
    last_name = last_name.upper() if isinstance(last_name, str) else last_name
    mi = mi.upper() if isinstance(mi, str) else mi
    suffix = suffix.upper() if isinstance(suffix, str) else suffix

    resolved_personnel_id = preserved_personnel_id or resolve_personnel_id(
        db,
        personnel_id,
        first_name,
        last_name,
        suffix,
        unit,
        mi,
    )

    # Parse date early to enforce monthly uniqueness
    parsed_date_check = parse_date_taken(date_taken) if date_taken else None
    parsed_date_check = parsed_date_check or datetime.utcnow()

    # Mark any previously-latest records for this personnel/name as NOT latest
    db.query(models.BMIRecord).filter(
        models.BMIRecord.name.ilike(name),
        models.BMIRecord.is_latest == True,
    ).update({models.BMIRecord.is_latest: False}, synchronize_session=False)
    db.commit()
    
    # Use explicit name parts when provided (more reliable than parsing full name)
    parts = name.strip().split() if name else []
    first_name_for_path = (first_name or '').strip() or (parts[0] if parts else '')
    last_name_for_path = (last_name or '').strip() or (parts[-1] if len(parts) > 1 else '')

    safe_unit_folder = safe_path_component(unit, 'UNIT').upper()
    safe_first = safe_path_component(first_name_for_path, 'FIRST').upper()
    safe_last = safe_path_component(last_name_for_path, 'LAST').upper()
    safe_mi = safe_path_component(mi, 'MI').upper() if mi and mi.strip() else ''
    folder_abs = uploads_abs('bmi', safe_unit_folder, bmi_folder_name(safe_first, safe_last, safe_mi))
    folder_rel = uploads_rel('bmi', safe_unit_folder, bmi_folder_name(safe_first, safe_last, safe_mi))
    os.makedirs(folder_abs, exist_ok=True)
    base_name = f"{safe_first}_{safe_mi}_{safe_last}" if safe_mi and safe_last else (f"{safe_first}_{safe_last}" if safe_last else safe_first)
    
    # Determine photo paths for the NEW record
    # Use new photos if uploaded, otherwise reuse old photo paths
    front_photo_path = preserved_old_front
    left_photo_path = preserved_old_left
    right_photo_path = preserved_old_right
    
    # Handle new front photo
    if photo_front and photo_front.filename:
        front_name = f"BMI_{base_name}_front.jpg"
        front_path_abs = os.path.join(str(folder_abs), front_name)
        with open(front_path_abs, 'wb') as f:
            f.write(await photo_front.read())
        front_photo_path = f"{folder_rel}/{front_name}".replace("\\", "/")
    
    # Handle new left photo
    if photo_left and photo_left.filename:
        left_name = f"BMI_{base_name}_left.jpg"
        left_path_abs = os.path.join(str(folder_abs), left_name)
        with open(left_path_abs, 'wb') as f:
            f.write(await photo_left.read())
        left_photo_path = f"{folder_rel}/{left_name}".replace("\\", "/")
    
    # Handle new right photo
    if photo_right and photo_right.filename:
        right_name = f"BMI_{base_name}_right.jpg"
        right_path_abs = os.path.join(str(folder_abs), right_name)
        with open(right_path_abs, 'wb') as f:
            f.write(await photo_right.read())
        right_photo_path = f"{folder_rel}/{right_name}".replace("\\", "/")
    
    # Parse date
    parsed_date = parse_date_taken(date_taken) if date_taken else None
    parsed_date = parsed_date or datetime.utcnow()
    
    # Calculate new BMI and classification
    bmi_value = compute_bmi(weight_kg, height_cm)
    pnp_classification = classify_pnp_bmi(bmi_value, age)
    
    # Create the NEW record using ORM to ensure DB compatibility
    try:
        new_record = models.BMIRecord(
            personnel_id=resolved_personnel_id,
            rank=rank,
            name=name,
            unit=unit,
            age=age,
            sex=sex,
            height_cm=height_cm,
            weight_kg=weight_kg,
            waist_cm=waist_cm,
            hip_cm=hip_cm,
            wrist_cm=wrist_cm,
            date_taken=parsed_date,
            bmi=bmi_value,
            classification=pnp_classification,
            result=f"{bmi_value:.2f}",
            photo_front=front_photo_path,
            photo_left=left_photo_path,
            photo_right=right_photo_path,
            status=status,
            status_custom=status_custom,
            is_latest=True,
        )
        db.add(new_record)
        db.commit()
        db.refresh(new_record)
    except Exception as e:
        logger.exception('Failed to create new BMI record during update')
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    # If we linked to an existing Personnel and MI was provided, persist it if not already set
    if resolved_personnel_id and mi and (mi.strip()):
        try:
            person = db.query(models.Personnel).filter(models.Personnel.id == resolved_personnel_id).first()
            if person and (not getattr(person, 'mi', None) or not str(person.mi).strip()):
                person.mi = mi
                db.add(person)
                db.commit()
        except Exception:
            pass

    # Upsert BMI history snapshot for the month (used by the 14-month PDF timeline)
    if resolved_personnel_id:
        history = db.query(BMIHistory).filter(
            BMIHistory.personnel_id == resolved_personnel_id,
            BMIHistory.month == parsed_date.month,
            BMIHistory.year == parsed_date.year,
        ).first()
        if history:
            history.weight = float(weight_kg)
            history.bmi_result = float(bmi_value)
            history.bmi_classification = pnp_classification
            history.waist = waist_cm
            history.hip = hip_cm
            history.wrist = wrist_cm
            history.created_at = parsed_date
        else:
            history = BMIHistory(
                personnel_id=resolved_personnel_id,
                month=parsed_date.month,
                year=parsed_date.year,
                weight=float(weight_kg),
                bmi_result=float(bmi_value),
                bmi_classification=pnp_classification,
                waist=waist_cm,
                hip=hip_cm,
                wrist=wrist_cm,
                created_at=parsed_date,
            )
            db.add(history)
        db.commit()
    
    # Return the NEW record (which is now the latest)
    return new_record


@router.delete('/{record_id}')
def delete_bmi_record(record_id: int, db: Session = Depends(get_db)):
    record = db.query(models.BMIRecord).filter(models.BMIRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail='BMI record not found')

    is_latest = bool(record.is_latest)
    personnel_id = record.personnel_id
    name = (record.name or '').strip()

    photo_paths = [
        resolve_upload_path(record.photo_front),
        resolve_upload_path(record.photo_left),
        resolve_upload_path(record.photo_right),
    ]

    db.delete(record)
    db.commit()

    if is_latest:
        candidate_query = db.query(models.BMIRecord)
        if personnel_id:
            candidate_query = candidate_query.filter(models.BMIRecord.personnel_id == personnel_id)
        elif name:
            candidate_query = candidate_query.filter(models.BMIRecord.name.ilike(name))

        latest_remaining = candidate_query.order_by(models.BMIRecord.date_taken.desc(), models.BMIRecord.id.desc()).first()
        if latest_remaining:
            latest_remaining.is_latest = True
            db.commit()

    for photo_path in photo_paths:
        try:
            if photo_path and os.path.exists(photo_path):
                os.remove(photo_path)
        except OSError:
            pass

    return {'message': 'BMI record deleted successfully'}


@router.get('/record/{record_id}', response_model=schemas.BMISchema)
def get_bmi_record(record_id: int, db: Session = Depends(get_db)):
    """
    Get a single BMI record by ID.
    Used to prefill the update form.
    """
    record = db.query(models.BMIRecord).filter(models.BMIRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="BMI record not found")
    return record
